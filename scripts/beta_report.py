"""
Beta Report — verifica el cálculo ARR sin Salesforce ni frontend.

Lee el Excel, ejecuta el pipeline completo (o usa datos ya importados),
y muestra un informe legible en terminal para que el CFO pueda validar
los números antes de conectar Salesforce.

Uso:
    # Primer uso: importa el Excel y genera el informe
    python scripts/beta_report.py --reimport

    # Usos siguientes: usa el snapshot ya importado
    python scripts/beta_report.py

    # Informe de un mes concreto
    python scripts/beta_report.py --month 2025-01

    # Exportar el informe a fichero
    python scripts/beta_report.py --output informe_beta.txt

Requiere:
    docker-compose up -d
    alembic upgrade head
"""

import argparse
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

EXCEL_DEFAULT = Path(__file__).resolve().parents[1] / "data_samples" / "raw_excel" / "ARR Oportunidad.xlsx"
SEP = "=" * 70
SEP2 = "-" * 70


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_eur(value: Decimal) -> str:
    """Format Decimal as € with thousands separator."""
    try:
        n = int(round(float(value)))
        return f"€{n:,}".replace(",", ".")
    except Exception:
        return str(value)


def _fmt_pct(value: Optional[float]) -> str:
    if value is None:
        return "  —   "
    sign = "+" if value >= 0 else ""
    return f"{sign}{value:.1f}%"


def _fmt_mom(value: Optional[Decimal]) -> str:
    if value is None:
        return "      —      "
    sign = "+" if value >= 0 else ""
    return f"{sign}{_fmt_eur(value)}"


# ---------------------------------------------------------------------------
# Import from Excel (calls import_excel_data.py logic)
# ---------------------------------------------------------------------------

def run_import(excel_path: str):
    """Full import: productos + consultores + snapshot con cálculo."""
    import importlib.util, os
    script = Path(__file__).resolve().parent / "import_excel_data.py"
    spec = importlib.util.spec_from_file_location("import_excel_data", script)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)

    import openpyxl
    print(f"Cargando Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    products = m.load_product_classifications(wb)
    countries = m.load_consultant_countries(wb)
    rows = list(m.load_opos_rows(wb))
    print(f"  {len(rows)} oportunidades leídas, {len(products)} productos, {len(countries)} consultores")

    from app.backend.db.connection import SessionLocal
    session = SessionLocal()
    try:
        m.upsert_product_classifications(session, products)
        m.upsert_consultant_countries(session, countries)
        snapshot = m.create_snapshot(session)
        raw_items = m.insert_raw_items(session, snapshot.id, rows)
        m.run_calculation_and_store(session, snapshot, raw_items)
        snapshot.status = "completed"
        session.commit()
        print(f"  Snapshot creado: {snapshot.id}\n")
        return str(snapshot.id)
    except Exception as e:
        session.rollback()
        raise
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Data queries (direct DB, no API needed)
# ---------------------------------------------------------------------------

def get_snapshot(session, snapshot_id=None):
    from app.backend.db.models import Snapshot
    if snapshot_id:
        return session.query(Snapshot).filter_by(id=snapshot_id).first()
    return (
        session.query(Snapshot)
        .filter(Snapshot.status.in_(["completed", "success"]))
        .order_by(Snapshot.created_at.desc())
        .first()
    )


def get_monthly_summary(session, snap_id):
    from app.backend.db.models import ARRMonthlySummary
    rows = (
        session.query(ARRMonthlySummary)
        .filter_by(snapshot_id=snap_id)
        .order_by(ARRMonthlySummary.month)
        .all()
    )
    # Group: {month → {product_type → arr}}
    summary = {}
    for r in rows:
        summary.setdefault(r.month, {})[r.product_type] = Decimal(str(r.arr_value))
    return summary


def get_consultant_arr(session, snap_id, target_month: date):
    from sqlalchemy import func
    from app.backend.db.models import ARRLineItem, RawOpportunityLineItem
    rows = (
        session.query(
            RawOpportunityLineItem.opportunity_owner,
            ARRLineItem.consultant_country,
            ARRLineItem.product_type,
            func.sum(ARRLineItem.annualized_value).label("arr"),
        )
        .join(ARRLineItem, ARRLineItem.raw_line_item_id == RawOpportunityLineItem.id)
        .filter(
            ARRLineItem.snapshot_id == snap_id,
            ARRLineItem.is_saas == True,
            ARRLineItem.start_month <= target_month,
            ARRLineItem.end_month_normalized >= target_month,
        )
        .group_by(
            RawOpportunityLineItem.opportunity_owner,
            ARRLineItem.consultant_country,
            ARRLineItem.product_type,
        )
        .all()
    )
    # {consultant → {total, country, by_type}}
    result = {}
    for owner, country, ptype, arr in rows:
        name = owner or "Sin consultor"
        if name not in result:
            result[name] = {"country": country or "—", "total": Decimal("0"), "by_type": {}}
        result[name]["total"] += Decimal(str(arr))
        result[name]["by_type"][ptype or "Unknown"] = (
            result[name]["by_type"].get(ptype or "Unknown", Decimal("0")) + Decimal(str(arr))
        )
    return result


def get_alerts(session, snap_id):
    from app.backend.db.models import SnapshotAlert
    return (
        session.query(SnapshotAlert)
        .filter_by(snapshot_id=snap_id, reviewed=False)
        .order_by(SnapshotAlert.severity, SnapshotAlert.alert_type)
        .all()
    )


def get_top_opportunities(session, snap_id, target_month: date, top_n=10):
    from sqlalchemy import func
    from app.backend.db.models import ARRLineItem, RawOpportunityLineItem
    rows = (
        session.query(
            RawOpportunityLineItem.opportunity_name,
            RawOpportunityLineItem.account_name,
            RawOpportunityLineItem.opportunity_owner,
            ARRLineItem.product_type,
            func.sum(ARRLineItem.annualized_value).label("arr"),
        )
        .join(ARRLineItem, ARRLineItem.raw_line_item_id == RawOpportunityLineItem.id)
        .filter(
            ARRLineItem.snapshot_id == snap_id,
            ARRLineItem.is_saas == True,
            ARRLineItem.start_month <= target_month,
            ARRLineItem.end_month_normalized >= target_month,
        )
        .group_by(
            RawOpportunityLineItem.opportunity_name,
            RawOpportunityLineItem.account_name,
            RawOpportunityLineItem.opportunity_owner,
            ARRLineItem.product_type,
        )
        .order_by(func.sum(ARRLineItem.annualized_value).desc())
        .limit(top_n)
        .all()
    )
    return rows


def get_excel_monthly_arr(excel_path: str, product_classifications: dict) -> dict:
    """
    Re-calcula el ARR del Excel mes a mes para comparar.
    Devuelve {month: total_arr}.
    """
    import openpyxl
    from datetime import timedelta

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Opos con Productos"]

    def _pd(v):
        if v is None: return None
        if isinstance(v, datetime): return v.date()
        if isinstance(v, date): return v
        s = str(v).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try: return datetime.strptime(s, fmt).date()
            except ValueError: pass
        return None

    def _dec(v):
        if v is None: return None
        try: return Decimal(str(v))
        except: return None

    # {month → arr}
    monthly = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        product_name = str(row[9]).strip() if row[9] else None
        if not product_name: continue
        product_type = product_classifications.get(product_name)
        if not product_type or not product_type.startswith("SaaS"): continue

        close_date = _pd(row[6])
        unit_price = _dec(row[10])
        quantity = _dec(row[15])
        if not close_date or unit_price is None or quantity is None: continue

        start = _pd(row[11]) or close_date
        end = _pd(row[12]) or (start + timedelta(days=365))
        if start > end: continue

        real_price = quantity * unit_price
        start_month = start.replace(day=1)
        raw_days = (end - start).days
        if raw_days <= 0: raw_days = 30
        end_month_normalized = start_month + timedelta(days=raw_days - 1)
        service_days = (end_month_normalized - start_month).days
        if service_days <= 0: service_days = 30

        annualized = real_price / Decimal(str(service_days)) * Decimal("365")

        # Distribute across active months
        m = start_month
        while m <= end_month_normalized:
            monthly[m] = monthly.get(m, Decimal("0")) + annualized
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)

    return monthly


# ---------------------------------------------------------------------------
# Report sections
# ---------------------------------------------------------------------------

def print_header(snap, out):
    out(SEP)
    out("  ARR COMMAND CENTER — INFORME DE VALIDACIÓN BETA")
    out(SEP)
    out(f"  Snapshot : {snap.id}")
    out(f"  Tipo     : {snap.sync_type}")
    out(f"  Creado   : {snap.created_at.strftime('%d %b %Y %H:%M') if snap.created_at else '—'}")
    out(f"  Registros: {snap.sf_records_processed or '—'}")
    out(f"  Alertas  : {snap.alerts_count or 0}")
    out(SEP)


def print_monthly_arr(monthly_summary, out, target_month=None):
    out("\n1. ARR MENSUAL POR LÍNEA DE NEGOCIO")
    out(SEP2)

    sorted_months = sorted(monthly_summary.keys())
    if not sorted_months:
        out("  (sin datos)")
        return

    # Collect all product types
    all_types = sorted({pt for m in monthly_summary.values() for pt in m})

    # Header
    header = f"  {'Mes':<12}"
    for pt in all_types:
        short = pt.replace("SaaS ", "").replace("isEazy ", "")[:10]
        header += f"  {short:>12}"
    header += f"  {'TOTAL':>12}  {'MoM':>9}"
    out(header)
    out("  " + "-" * (len(header) - 2))

    prev_total = None
    for month in sorted_months:
        if target_month and month != target_month:
            pass  # still print all, but highlight target
        by_type = monthly_summary[month]
        total = sum(by_type.values(), Decimal("0"))
        mom_pct = (
            float((total - prev_total) / prev_total * 100)
            if prev_total and prev_total != 0 else None
        )
        row = f"  {month.strftime('%b %Y'):<12}"
        for pt in all_types:
            v = by_type.get(pt, Decimal("0"))
            row += f"  {_fmt_eur(v):>12}"
        row += f"  {_fmt_eur(total):>12}  {_fmt_pct(mom_pct):>9}"
        marker = " ◄" if target_month and month == target_month else ""
        out(row + marker)
        prev_total = total


def print_comparison_vs_excel(monthly_summary, excel_monthly, out):
    out("\n2. COMPARACIÓN APP vs EXCEL (solo meses con datos en ambos)")
    out(SEP2)
    out(f"  {'Mes':<12}  {'App (SaaS)':>14}  {'Excel (SaaS)':>14}  {'Diferencia':>12}  {'Estado':>8}")
    out("  " + "-" * 65)

    all_months = sorted(set(monthly_summary.keys()) | set(excel_monthly.keys()))
    max_diff = Decimal("0")
    fails = 0

    for month in all_months:
        app_total = sum(monthly_summary.get(month, {}).values(), Decimal("0"))
        excel_total = excel_monthly.get(month, Decimal("0"))
        diff = app_total - excel_total
        pct = float(diff / excel_total * 100) if excel_total != 0 else 0.0
        status = "OK" if abs(pct) < 1.0 else "REVISAR"
        if status == "REVISAR":
            fails += 1
        max_diff = max(max_diff, abs(diff))
        out(
            f"  {month.strftime('%b %Y'):<12}  {_fmt_eur(app_total):>14}  "
            f"{_fmt_eur(excel_total):>14}  {_fmt_mom(diff):>12}  {status:>8}"
        )

    out(SEP2)
    if fails == 0:
        out(f"  RESULTADO: OK — diferencia máxima {_fmt_eur(max_diff)} (< 1% en todos los meses)")
    else:
        out(f"  RESULTADO: {fails} mes(es) con diferencia > 1% — revisar manualmente")


def print_consultant_table(consultant_data, target_month, out, top_n=15):
    out(f"\n3. ARR POR CONSULTOR — {target_month.strftime('%B %Y').upper()}")
    out(SEP2)
    out(f"  {'Consultor':<25}  {'País':<10}  {'ARR Total':>12}  {'Línea principal':<20}")
    out("  " + "-" * 75)

    sorted_consultants = sorted(
        consultant_data.items(), key=lambda x: x[1]["total"], reverse=True
    )[:top_n]

    for name, data in sorted_consultants:
        top_line = max(data["by_type"], key=data["by_type"].get) if data["by_type"] else "—"
        top_line_short = top_line.replace("SaaS ", "")
        out(
            f"  {name[:25]:<25}  {data['country'][:10]:<10}  "
            f"{_fmt_eur(data['total']):>12}  {top_line_short:<20}"
        )


def print_top_opportunities(top_opps, target_month, out):
    out(f"\n4. TOP 10 OPORTUNIDADES POR ARR — {target_month.strftime('%B %Y').upper()}")
    out(SEP2)
    out(f"  {'Oportunidad':<35}  {'Consultor':<15}  {'Tipo':<15}  {'ARR':>12}")
    out("  " + "-" * 82)

    for opp_name, account, owner, ptype, arr in top_opps:
        name_short = (opp_name or account or "—")[:35]
        owner_short = (owner or "—")[:15]
        type_short = (ptype or "—").replace("SaaS ", "")[:15]
        out(f"  {name_short:<35}  {owner_short:<15}  {type_short:<15}  {_fmt_eur(Decimal(str(arr))):>12}")


def print_alerts(alerts, out):
    out(f"\n5. ALERTAS DE CALIDAD DE DATOS ({len(alerts)} sin revisar)")
    out(SEP2)

    if not alerts:
        out("  Sin alertas pendientes.")
        return

    by_type = {}
    for a in alerts:
        by_type.setdefault(a.alert_type, []).append(a)

    for alert_type, items in sorted(by_type.items()):
        severity_icon = "🔴" if items[0].severity == "error" else "🟡"
        out(f"\n  {severity_icon} {alert_type} ({len(items)} casos)")
        for a in items[:5]:
            opp = a.opportunity_name or "—"
            prod = a.product_name or "—"
            out(f"     · {opp[:50]} | {prod[:25]}")
        if len(items) > 5:
            out(f"     ... y {len(items)-5} más")


def print_footer(snap, out):
    out(f"\n{SEP}")
    out("  CÓMO INTERPRETAR ESTE INFORME")
    out(SEP2)
    out("  • Sección 2: diferencias < 1% son aceptables (redondeos).")
    out("  • Sección 5 errores 🔴: productos sin clasificar excluidos del ARR.")
    out("    → Añadirlos en /api/config/products y volver a ejecutar --reimport.")
    out("  • Alertas 🟡: datos incompletos en SF (fechas vacías, duraciones raras).")
    out("    → Se usan fallbacks; verificar manualmente si el ARR parece correcto.")
    out(f"\n  Para iniciar la API:  uvicorn app.backend.main:app --reload --port 8000")
    out(f"  Docs interactivas:   http://localhost:8000/docs")
    out(SEP)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Beta validation report — no Salesforce needed")
    parser.add_argument("--excel", default=str(EXCEL_DEFAULT), help="Ruta al Excel")
    parser.add_argument("--env", default=".env", help="Ruta al .env")
    parser.add_argument("--reimport", action="store_true", help="Reimportar desde el Excel antes de generar el informe")
    parser.add_argument("--snapshot-id", default=None, help="UUID del snapshot a analizar")
    parser.add_argument("--month", default=None, help="Mes objetivo (YYYY-MM). Por defecto: último mes con datos")
    parser.add_argument("--output", default=None, help="Escribir informe a fichero además de terminal")
    args = parser.parse_args()

    from dotenv import load_dotenv
    load_dotenv(args.env)

    # Collect output lines
    lines = []
    def out(text=""):
        print(text)
        lines.append(text)

    if args.reimport:
        run_import(args.excel)

    from app.backend.db.connection import SessionLocal
    from app.backend.db.models import ProductClassification
    session = SessionLocal()
    try:
        snap = get_snapshot(session, args.snapshot_id)
        if not snap:
            print("ERROR: No se encontró ningún snapshot completado.")
            print("       Ejecuta: python scripts/beta_report.py --reimport")
            sys.exit(1)

        monthly_summary = get_monthly_summary(session, snap.id)
        if not monthly_summary:
            print("ERROR: El snapshot no tiene datos de resumen mensual.")
            print("       Prueba con --reimport para regenerar.")
            sys.exit(1)

        # Determine target month
        sorted_months = sorted(monthly_summary.keys())
        if args.month:
            y, m = map(int, args.month.split("-"))
            target_month = date(y, m, 1)
        else:
            target_month = sorted_months[-1]

        consultant_data = get_consultant_arr(session, snap.id, target_month)
        top_opps = get_top_opportunities(session, snap.id, target_month)
        alerts = get_alerts(session, snap.id)

        # Load product classifications for Excel comparison
        product_map = {
            p.product_name: p.product_type
            for p in session.query(ProductClassification).all()
        }

    finally:
        session.close()

    # Build Excel comparison
    try:
        excel_monthly = get_excel_monthly_arr(args.excel, product_map)
    except Exception as e:
        excel_monthly = {}
        print(f"  Aviso: no se pudo leer el Excel para comparación ({e})")

    # Generate report
    print_header(snap, out)
    print_monthly_arr(monthly_summary, out, target_month)
    if excel_monthly:
        print_comparison_vs_excel(monthly_summary, excel_monthly, out)
    print_consultant_table(consultant_data, target_month, out)
    print_top_opportunities(top_opps, target_month, out)
    print_alerts(alerts, out)
    print_footer(snap, out)

    # Write to file if requested
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        print(f"\nInforme guardado en: {args.output}")


if __name__ == "__main__":
    main()
