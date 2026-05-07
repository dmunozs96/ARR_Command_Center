"""
Compare ARR monthly totals: Python calculator vs Excel Resumen.

Reads Experimento_Claude_ARR.xlsx and:
  1. Builds "Excel ground truth" from Opos cols (col 20 = product_type,
     col 24 = start_month, col 25 = end_month_normalized, col 35 = annualized_value).
  2. Runs our Python ARRCalculator on the same raw input rows + master tables
     from the same file.
  3. Compares both against the Resumen sheet monthly values.

Usage:
    python scripts/compare_monthly_arr.py [--excel PATH] [--year YYYY]
"""

import argparse
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from app.backend.core.arr_calculator import ARRCalculator, RawLineItem, _last_day_of_month
from app.backend.core.excel_importer import (
    _normalize_key,
    load_consultant_countries,
    load_opos_rows,
    load_product_classifications,
)

EXCEL_DEFAULT = (
    Path(__file__).resolve().parents[1]
    / "data_samples"
    / "expected_outputs"
    / "Experimento_Claude_ARR.xlsx"
)
TOLERANCE = Decimal("1.00")  # 1€ per month/product_type cell

# ── Resumen row → our product_type ──────────────────────────────────────────
# Identified by the Excel label in col 3 of the Resumen sheet.
RESUMEN_LABEL_TO_PRODUCT_TYPE = {
    "all in one": "SaaS AIO",
    "iseazy author offline": "SaaS Author",
    "iseazy author online": "Author Online",
    "iseazy engage": "SaaS Engage",
    "iseazy lms": "SaaS LMS",
    "all in one iseazy lms": None,  # combined row – skip
    "iseazy skills": "SaaS Skills",
    "total arr compania": None,     # total row – check separately
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _dt_to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------------------
# 1. Read "Excel ground truth" directly from Opos cols
# ---------------------------------------------------------------------------

def read_excel_ground_truth(ws) -> dict[date, dict[str, Decimal]]:
    """
    Build monthly summaries by reading pre-calculated columns directly:
      col 20 = Tipo de Producto Correcto
      col 24 = inicio mes  (start_month)
      col 25 = fin mes     (end_month_normalized)
      col 35 = servicio anualizado

    Only rows with a SaaS-type product are included.
    """
    line_items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if len(row) < 36:
            continue

        product_type = row[20]
        if not product_type or not str(product_type).startswith("SaaS"):
            continue

        start_month = _dt_to_date(row[24])
        end_month = _dt_to_date(row[25])
        annualized = row[35]

        if start_month is None or end_month is None or annualized is None:
            continue
        try:
            annualized = Decimal(str(annualized))
        except Exception:
            continue

        if annualized <= 0:
            continue

        line_items.append((str(product_type), start_month, end_month, annualized))

    # Aggregate by month / product_type
    months_set: set[date] = set()
    for _, sm, em, _ in line_items:
        m = sm
        while m <= em:
            months_set.add(m)
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)

    result: dict[date, dict[str, Decimal]] = {}
    for month in months_set:
        month_end = _last_day_of_month(month)
        by_type: dict[str, Decimal] = {}
        for pt, sm, em, arr in line_items:
            if sm <= month_end and em >= month:
                by_type[pt] = by_type.get(pt, Decimal("0")) + arr
        result[month] = by_type

    return result


# ---------------------------------------------------------------------------
# 2. Read Resumen expected values
# ---------------------------------------------------------------------------

def read_resumen(wb) -> dict[date, dict[str, Decimal]]:
    """
    Parse the Resumen sheet — "Fecha de Inicio" section only (rows 8-17).

    The sheet has two calculation modes:
      row 8  header = "Fecha de Inicio"  (rows 8-17)  ← our methodology
      row 78 header = "Fecha de Cierre"  (rows 78-87) ← alternative, skip

    Layout:
      row 4  = month dates (first day of month), col 5 onwards;
               col 17 is a None separator between years.
    """
    ws = wb["Resumen"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Parse month columns from row 4 (index 3)
    month_row = all_rows[3]
    col_to_month: dict[int, date] = {}
    for col_idx, val in enumerate(month_row):
        d = _dt_to_date(val)
        if d is not None:
            col_to_month[col_idx] = d.replace(day=1)

    if not col_to_month:
        raise ValueError("Could not parse month headers from Resumen row 4.")

    result: dict[date, dict[str, Decimal]] = {m: {} for m in col_to_month.values()}

    # Parse product rows — stop when we hit "Fecha de Cierre" header (row 78)
    in_fecha_inicio_section = False
    for row in all_rows[7:]:  # data starts row 8 (index 7)
        # Detect section headers in col 0
        section_label = _normalize_key(str(row[0])) if row and row[0] else ""
        if "fecha de inicio" in section_label:
            in_fecha_inicio_section = True
        elif "fecha de cierre" in section_label:
            break  # stop — rest is alternative methodology

        sublabel_raw = row[3] if len(row) > 3 else None
        if sublabel_raw is None:
            continue
        sublabel = _normalize_key(str(sublabel_raw))
        product_type = RESUMEN_LABEL_TO_PRODUCT_TYPE.get(sublabel)
        if product_type is None:
            continue  # skip combined / total / variation rows

        for col_idx, month in col_to_month.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            try:
                result[month][product_type] = Decimal(str(val))
            except Exception:
                pass

    return result


# ---------------------------------------------------------------------------
# 3. Run Python ARRCalculator
# ---------------------------------------------------------------------------

def run_python_calculator(wb) -> dict[date, dict[str, Decimal]]:
    products_raw = load_product_classifications(wb)
    countries_raw = load_consultant_countries(wb)
    rows = list(load_opos_rows(wb))

    # Build flat lookup: plain + compound keys → product_type string
    products: dict[str, str] = {}
    for name, info in products_raw.items():
        pt = info.get("product_type") if isinstance(info, dict) else info
        if pt:
            products[name] = pt

    countries: dict[str, str] = {k: v for k, v in countries_raw.items()}

    calc = ARRCalculator(products, countries)

    domain_items = []
    for row in rows:
        domain_items.append(RawLineItem(
            sf_opportunity_id=row["sf_opportunity_id"],
            sf_line_item_id=row["sf_line_item_id"],
            opportunity_name=row["opportunity_name"] or "",
            account_name=row["account_name"] or "",
            opportunity_owner=row["opportunity_owner"] or "",
            opportunity_type=row["opportunity_type"] or "",
            channel_type=row["channel_type"] or "",
            close_date=row["close_date"],
            product_name=row["product_name"],
            unit_price=row["unit_price"],
            quantity=row["quantity"],
            subscription_start_date=row["subscription_start_date"],
            subscription_end_date=row["subscription_end_date"],
            licence_period_months=row["licence_period_months"],
            business_line=row["business_line"],
            opportunity_amount=row["opportunity_amount"],
            product_code=row["product_code"],
        ))

    snapshot = calc.process_all(domain_items)

    # Collect all months
    months_set: set[date] = set()
    for item in snapshot.saas_items():
        m = item.start_month
        while m <= item.end_month_normalized:
            months_set.add(m)
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)

    result: dict[date, dict[str, Decimal]] = {}
    for month in months_set:
        month_end = _last_day_of_month(month)
        by_type: dict[str, Decimal] = {}
        for item in snapshot.saas_items():
            if item.start_month <= month_end and item.end_month_normalized >= month:
                pt = item.product_type or "UNCLASSIFIED"
                by_type[pt] = by_type.get(pt, Decimal("0")) + item.annualized_value
        result[month] = by_type

    return result


# ---------------------------------------------------------------------------
# 4. Compare and report
# ---------------------------------------------------------------------------

def compare_and_report(
    excel_gt: dict,
    python_calc: dict,
    resumen: dict,
    year_filter: Optional[int],
    show_ok: bool,
):
    all_months = sorted(set(excel_gt) | set(python_calc) | set(resumen))
    if year_filter:
        all_months = [m for m in all_months if m.year == year_filter]

    all_product_types = sorted(
        set(pt for by_type in excel_gt.values() for pt in by_type)
        | set(pt for by_type in python_calc.values() for pt in by_type)
        | set(pt for by_type in resumen.values() for pt in by_type)
    )

    header = f"{'Month':<12} {'ProductType':<20} {'Excel_Opos':>14} {'Python_Calc':>14} {'Resumen':>14} {'Py-Excel':>12} {'Py-Resumen':>12}"
    separator = "-" * len(header)

    # Author Online comes from Stripe, not from our Salesforce calculator.
    # It appears in the Resumen but will always be 0 in Python_Calc and Excel_GT.
    # We show it as informational but don't count it as a mismatch.
    STRIPE_ONLY_TYPES = {"Author Online"}

    print(header)
    print(separator)

    total_gt = Decimal("0")
    total_py = Decimal("0")
    mismatches_py_excel = 0
    mismatches_py_resumen = 0

    for month in all_months:
        gt_by_type = excel_gt.get(month, {})
        py_by_type = python_calc.get(month, {})
        res_by_type = resumen.get(month, {})

        for pt in all_product_types:
            gt_val = gt_by_type.get(pt, Decimal("0"))
            py_val = py_by_type.get(pt, Decimal("0"))
            res_val = res_by_type.get(pt, Decimal("0"))

            diff_py_excel = py_val - gt_val
            diff_py_resumen = py_val - res_val

            is_stripe_only = pt in STRIPE_ONLY_TYPES
            is_mismatch_excel = (not is_stripe_only) and abs(diff_py_excel) >= TOLERANCE
            is_mismatch_resumen = (not is_stripe_only) and abs(diff_py_resumen) >= TOLERANCE

            if is_mismatch_excel:
                mismatches_py_excel += 1
            if is_mismatch_resumen:
                mismatches_py_resumen += 1

            if show_ok or is_mismatch_excel or is_mismatch_resumen:
                flag = ""
                if is_stripe_only:
                    flag = " [Stripe/manual]"
                elif is_mismatch_excel and is_mismatch_resumen:
                    flag = " ** BOTH"
                elif is_mismatch_excel:
                    flag = " ** PY!=EXCEL"
                elif is_mismatch_resumen:
                    flag = " ** PY!=RESUMEN"

                print(
                    f"{month.strftime('%Y-%m'):<12} {pt:<20} "
                    f"{float(gt_val):>14.2f} {float(py_val):>14.2f} "
                    f"{float(res_val):>14.2f} "
                    f"{float(diff_py_excel):>+12.2f} {float(diff_py_resumen):>+12.2f}"
                    f"{flag}"
                )

        month_gt = sum(gt_by_type.values(), Decimal("0"))
        month_py = sum(py_by_type.values(), Decimal("0"))

        total_gt += month_gt
        total_py += month_py

    print(separator)
    print(f"\nSummary (mismatches >= {TOLERANCE}€):")
    print(f"  Python vs Excel Opos cols : {mismatches_py_excel} cell(s)")
    print(f"  Python vs Resumen sheet   : {mismatches_py_resumen} cell(s)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Compare monthly ARR: Python vs Excel")
    parser.add_argument("--excel", default=str(EXCEL_DEFAULT))
    parser.add_argument("--year", type=int, default=None, help="Filter to a single year (e.g. 2025)")
    parser.add_argument("--show-ok", action="store_true", help="Show matching rows too")
    parser.add_argument("--product", default=None, help="Filter by product type (e.g. 'SaaS LMS')")
    args = parser.parse_args()

    print(f"Loading: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

    print("Reading Excel ground truth from Opos cols...")
    ws_opos = wb["Opos con Productos"]
    excel_gt = read_excel_ground_truth(ws_opos)

    print("Reading Resumen expected values...")
    resumen = read_resumen(wb)

    print("Running Python ARRCalculator...")
    python_calc = run_python_calculator(wb)

    print()

    if args.product:
        # Filter to one product type
        excel_gt = {m: {args.product: v.get(args.product, Decimal("0"))} for m, v in excel_gt.items()}
        python_calc = {m: {args.product: v.get(args.product, Decimal("0"))} for m, v in python_calc.items()}
        resumen = {m: {args.product: v.get(args.product, Decimal("0"))} for m, v in resumen.items()}

    compare_and_report(excel_gt, python_calc, resumen, args.year, args.show_ok)


if __name__ == "__main__":
    main()
