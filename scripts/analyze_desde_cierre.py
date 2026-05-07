"""
Analyze 'Fecha de Cierre' logic from the Excel Resumen sheet.

Compares:
  A) Excel Resumen "Fecha de Cierre" section (rows 78-87)
  B) Current backend from_close logic:   start=close_date, end=end_month_normalized (ORIGINAL)
  C) Hypothesis fix:                      start=close_date, end=close_date + service_days

Usage:
    python scripts/analyze_desde_cierre.py [--year YYYY]
"""

import argparse
import sys
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from app.backend.core.arr_calculator import _last_day_of_month
from app.backend.core.excel_importer import (
    _normalize_key,
    load_consultant_countries,
    load_opos_rows,
    load_product_classifications,
)
from app.backend.core.arr_calculator import ARRCalculator, RawLineItem

EXCEL_DEFAULT = (
    Path(__file__).resolve().parents[1]
    / "data_samples"
    / "expected_outputs"
    / "Experimento_Claude_ARR.xlsx"
)
TOLERANCE = Decimal("1.00")

RESUMEN_LABEL_TO_PRODUCT_TYPE = {
    "all in one": "SaaS AIO",
    "iseazy author offline": "SaaS Author",
    "iseazy author online": "Author Online",
    "iseazy engage": "SaaS Engage",
    "iseazy lms": "SaaS LMS",
    "all in one iseazy lms": None,
    "iseazy skills": "SaaS Skills",
    "total arr compania": None,
}


def _dt_to_date(v) -> Optional[date]:
    from datetime import datetime
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def read_resumen_cierre(wb) -> dict[date, dict[str, Decimal]]:
    """Parse the Resumen 'Fecha de Cierre' section (rows 78-87)."""
    ws = wb["Resumen"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Parse month columns from row 4 (index 3)
    month_row = all_rows[3]
    col_to_month: dict[int, date] = {}
    for col_idx, val in enumerate(month_row):
        d = _dt_to_date(val)
        if d is not None:
            col_to_month[col_idx] = d.replace(day=1)

    result: dict[date, dict[str, Decimal]] = {m: {} for m in col_to_month.values()}

    in_cierre_section = False
    for row in all_rows[7:]:
        section_label = _normalize_key(str(row[0])) if row and row[0] else ""
        if "fecha de cierre" in section_label:
            in_cierre_section = True
            continue
        if in_cierre_section and "fecha de" in section_label and "cierre" not in section_label:
            break  # another section starts

        if not in_cierre_section:
            continue

        sublabel_raw = row[3] if len(row) > 3 else None
        if sublabel_raw is None:
            continue
        sublabel = _normalize_key(str(sublabel_raw))
        product_type = RESUMEN_LABEL_TO_PRODUCT_TYPE.get(sublabel)
        if product_type is None:
            continue

        for col_idx, month in col_to_month.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            try:
                result[month][product_type] = Decimal(str(val))
            except Exception:
                pass

    return result


def build_calc_items(wb):
    """Build ARR line items from the Excel data."""
    products_raw = load_product_classifications(wb)
    countries_raw = load_consultant_countries(wb)
    rows = list(load_opos_rows(wb))

    products: dict[str, str] = {}
    for name, info in products_raw.items():
        pt = info.get("product_type") if isinstance(info, dict) else info
        if pt:
            products[name] = pt

    countries: dict[str, str] = {k: v for k, v in countries_raw.items()}
    calc = ARRCalculator(products, countries)

    domain_items = []
    for row in rows:
        domain_items.append(RawLineItem(
            sf_opportunity_id=row["sf_opportunity_id"],
            sf_line_item_id=row["sf_line_item_id"],
            opportunity_name=row["opportunity_name"] or "",
            account_name=row["account_name"] or "",
            opportunity_owner=row["opportunity_owner"] or "",
            opportunity_type=row["opportunity_type"] or "",
            channel_type=row["channel_type"] or "",
            close_date=row["close_date"],
            product_name=row["product_name"],
            unit_price=row["unit_price"],
            quantity=row["quantity"],
            subscription_start_date=row["subscription_start_date"],
            subscription_end_date=row["subscription_end_date"],
            licence_period_months=row["licence_period_months"],
            business_line=row["business_line"],
            opportunity_amount=row["opportunity_amount"],
            product_code=row["product_code"],
        ))

    return calc.process_all(domain_items)


def aggregate_by_month(items, start_fn, end_fn) -> dict[date, dict[str, Decimal]]:
    """Aggregate ARR by month using custom start/end functions per item."""
    months_set: set[date] = set()
    tuples = []
    for item in items:
        s = start_fn(item)
        e = end_fn(item)
        if s > e:
            continue
        tuples.append((s, e, item.product_type, item.annualized_value))
        m = s
        while m <= e:
            months_set.add(m)
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)

    result: dict[date, dict[str, Decimal]] = {}
    for month in months_set:
        month_end = _last_day_of_month(month)
        by_type: dict[str, Decimal] = {}
        for s, e, pt, arr in tuples:
            if s <= month_end and e >= month:
                pt_key = pt or "UNCLASSIFIED"
                by_type[pt_key] = by_type.get(pt_key, Decimal("0")) + arr
        result[month] = by_type

    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=str(EXCEL_DEFAULT))
    parser.add_argument("--year", type=int, default=None)
    parser.add_argument("--product", default=None)
    args = parser.parse_args()

    print(f"Loading: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

    print("Reading Resumen 'Fecha de Cierre' section...")
    resumen_cierre = read_resumen_cierre(wb)

    # Check if we got any data
    total_cells = sum(len(v) for v in resumen_cierre.values())
    print(f"  Found {total_cells} data cells in Resumen Fecha de Cierre section")
    if total_cells == 0:
        print("  WARNING: No data found in Fecha de Cierre section. Dumping rows 75-95 for inspection:")
        ws = wb["Resumen"]
        all_rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(all_rows[74:95], start=75):
            if any(c is not None for c in row[:10]):
                print(f"    Row {i}: {row[:10]}")

    print("Building ARR line items from Excel data...")
    snapshot = build_calc_items(wb)
    saas_items = snapshot.saas_items()
    print(f"  {len(saas_items)} SaaS line items")

    # Method B: current backend logic (start=close_date, end=original end_month_normalized)
    method_b = aggregate_by_month(
        saas_items,
        start_fn=lambda i: i.raw.close_date.replace(day=1),
        end_fn=lambda i: i.end_month_normalized,
    )

    # Method C: fix hypothesis (start=close_date, end=close_start + service_days)
    method_c = aggregate_by_month(
        saas_items,
        start_fn=lambda i: i.raw.close_date.replace(day=1),
        end_fn=lambda i: i.raw.close_date.replace(day=1) + timedelta(days=i.service_days),
    )

    # Compare
    product_types = sorted({
        "SaaS AIO", "SaaS Author", "SaaS Engage", "SaaS LMS", "SaaS Skills"
    })
    if args.product:
        product_types = [args.product]

    all_months = sorted(set(resumen_cierre) | set(method_b) | set(method_c))
    if args.year:
        all_months = [m for m in all_months if m.year == args.year]

    header = f"{'Month':<12} {'ProductType':<20} {'Resumen_Cierre':>16} {'MethodB_Current':>16} {'MethodC_Fix':>12} {'B-Res':>10} {'C-Res':>10}"
    sep = "-" * len(header)
    print()
    print(header)
    print(sep)

    total_miss_b = 0
    total_miss_c = 0

    for m in all_months:
        res = resumen_cierre.get(m, {})
        b = method_b.get(m, {})
        c = method_c.get(m, {})

        for pt in product_types:
            res_v = res.get(pt, Decimal("0"))
            b_v = b.get(pt, Decimal("0"))
            c_v = c.get(pt, Decimal("0"))

            if res_v == 0 and b_v == 0 and c_v == 0:
                continue

            diff_b = b_v - res_v
            diff_c = c_v - res_v

            miss_b = abs(diff_b) >= TOLERANCE
            miss_c = abs(diff_c) >= TOLERANCE
            if miss_b:
                total_miss_b += 1
            if miss_c:
                total_miss_c += 1

            flag = ""
            if miss_b and not miss_c:
                flag = " [C FIXES]"
            elif miss_b and miss_c:
                flag = " [BOTH WRONG]"
            elif not miss_b and miss_c:
                flag = " [C BREAKS]"

            print(
                f"{m.strftime('%Y-%m'):<12} {pt:<20} "
                f"{float(res_v):>16.0f} {float(b_v):>16.0f} {float(c_v):>12.0f} "
                f"{float(diff_b):>+10.0f} {float(diff_c):>+10.0f}"
                f"{flag}"
            )

    print(sep)
    print(f"\nSummary (mismatches >= {TOLERANCE}€ vs Resumen Cierre):")
    print(f"  Method B (current backend from_close) : {total_miss_b} mismatches")
    print(f"  Method C (close_date + service_days)  : {total_miss_c} mismatches")


if __name__ == "__main__":
    main()
