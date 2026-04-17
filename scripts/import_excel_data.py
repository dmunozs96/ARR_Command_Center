"""
Import data from 'ARR Oportunidad.xlsx' into the database.

Creates a snapshot of type 'excel_import' with all line items from the
'Opos con Productos' sheet, plus product and consultant lookups from their
respective sheets.

Usage:
    cd ARR_Command_Center
    python scripts/import_excel_data.py [--excel PATH] [--env PATH]
"""

import argparse
import hashlib
import os
import sys
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

# Allow running from repo root
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

EXCEL_DEFAULT = Path(__file__).resolve().parents[1] / "data_samples" / "raw_excel" / "ARR Oportunidad.xlsx"


# ---------------------------------------------------------------------------
# Excel parsing helpers
# ---------------------------------------------------------------------------

def _parse_date_text(value) -> Optional[date]:
    """Parse dates stored as text (dd/mm/yyyy) or as Excel date numbers."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # Try Excel serial number
    try:
        n = float(s)
        return date(1899, 12, 30) + timedelta(days=int(n))
    except (ValueError, TypeError):
        return None


def _decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def _str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Read lookup tables from Excel
# ---------------------------------------------------------------------------

def load_product_classifications(wb) -> dict:
    """
    Returns {product_name: {"product_type": ..., "product_code": ...,
                             "category": ..., "subcategory": ..., "business_line": ...}}
    from 'Productos SF SAAS' sheet.

    Columns: B=product_name, C=product_code, E=business_line, F=category, G=subcategory, H=product_type
    """
    ws = wb["Productos SF SAAS"]
    result = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if row[1] is None:
            continue
        name = _str(row[1])
        if not name:
            continue
        result[name] = {
            "product_code": _str(row[2]),
            "business_line": _str(row[4]),
            "category": _str(row[5]),
            "subcategory": _str(row[6]),
            "product_type": _str(row[7]),
        }
    return result


def load_consultant_countries(wb) -> dict:
    """
    Returns {consultant_name: country} from 'País Consultor' sheet.
    Columns: C=consultant_name, D=country (header row at index 3 by content).
    """
    ws = wb["País Consultor"]
    result = {}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[2] == "Consultor":
                header_found = True
            continue
        name = _str(row[2])
        country = _str(row[3])
        if name and country:
            result[name] = country
    return result


# ---------------------------------------------------------------------------
# Read main data sheet
# ---------------------------------------------------------------------------

def load_opos_rows(wb):
    """
    Yield dicts for each data row in 'Opos con Productos'.
    Columns A–R (indices 0–17):
      0  Propietario de oportunidad (opportunity_owner)
      1  Nombre de la cuenta (account_name)
      2  Nombre de la oportunidad (opportunity_name)
      3  Tipo (opportunity_type)
      4  Tipo de oportunidad (channel_type)
      5  Importe (opportunity_amount)
      6  Fecha de cierre (close_date)  ← TEXT dd/mm/yyyy
      7  Fecha de creación (created_date_sf)
      8  Etapa (stage, always "Ganada")
      9  Nombre del producto (product_name)
      10 Precio de venta (unit_price)
      11 Subscription Start Date ← TEXT
      12 Subscription End Date ← TEXT
      13 Licence period (months)
      14 Línea de negocio (business_line)
      15 Cantidad (quantity)
      16 Product (product_code)
      17 Creado por
    """
    ws = wb["Opos con Productos"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if row[9] is None and row[2] is None:
            continue  # completely empty row

        close_date = _parse_date_text(row[6])
        if close_date is None:
            continue  # cannot process without close date

        product_name = _str(row[9])
        if not product_name:
            continue

        unit_price = _decimal(row[10]) or Decimal("0")
        quantity = _decimal(row[15]) or Decimal("1")

        # Synthetic SF IDs from row content (no real SF IDs in Excel)
        opp_key = f"{_str(row[2]) or ''}-{close_date}"
        sf_opportunity_id = "EXCEL_" + hashlib.md5(opp_key.encode()).hexdigest()[:12].upper()
        line_key = f"{opp_key}-{product_name}-{unit_price}-{quantity}"
        sf_line_item_id = "EXCL_" + hashlib.md5(line_key.encode()).hexdigest()[:13].upper()

        yield {
            "sf_opportunity_id": sf_opportunity_id,
            "sf_line_item_id": sf_line_item_id,
            "opportunity_owner": _str(row[0]) or "Unknown",
            "account_name": _str(row[1]),
            "opportunity_name": _str(row[2]),
            "opportunity_type": _str(row[3]),
            "channel_type": _str(row[4]),
            "opportunity_amount": _decimal(row[5]),
            "close_date": close_date,
            "product_name": product_name,
            "unit_price": unit_price,
            "subscription_start_date": _parse_date_text(row[11]),
            "subscription_end_date": _parse_date_text(row[12]),
            "licence_period_months": int(row[13]) if row[13] is not None else None,
            "business_line": _str(row[14]),
            "quantity": quantity,
            "product_code": _str(row[16]),
        }


# ---------------------------------------------------------------------------
# Database operations
# ---------------------------------------------------------------------------

def upsert_product_classifications(session, products: dict):
    from app.backend.db.models import ProductClassification
    count = 0
    for name, info in products.items():
        existing = session.query(ProductClassification).filter_by(product_name=name).first()
        if existing:
            existing.product_type = info["product_type"] or "UNCLASSIFIED"
            existing.product_code = info["product_code"]
            existing.category = info["category"]
            existing.subcategory = info["subcategory"]
            existing.business_line = info["business_line"]
        else:
            session.add(ProductClassification(
                product_name=name,
                product_type=info["product_type"] or "UNCLASSIFIED",
                product_code=info["product_code"],
                category=info["category"],
                subcategory=info["subcategory"],
                business_line=info["business_line"],
            ))
            count += 1
    session.flush()
    print(f"  Products seeded/updated: {len(products)} (new: {count})")


def upsert_consultant_countries(session, countries: dict):
    from app.backend.db.models import ConsultantCountry
    count = 0
    for name, country in countries.items():
        existing = session.query(ConsultantCountry).filter_by(consultant_name=name).first()
        if existing:
            existing.country = country
        else:
            session.add(ConsultantCountry(consultant_name=name, country=country))
            count += 1
    session.flush()
    print(f"  Consultants seeded/updated: {len(countries)} (new: {count})")


def create_snapshot(session) -> "Snapshot":
    from app.backend.db.models import Snapshot
    snap = Snapshot(
        id=uuid.uuid4(),
        sync_type="excel_import",
        triggered_by="import_excel_data.py",
        status="running",
        notes="Importado desde ARR Oportunidad.xlsx",
    )
    session.add(snap)
    session.flush()
    return snap


def insert_raw_items(session, snapshot_id, rows) -> list:
    from app.backend.db.models import RawOpportunityLineItem
    inserted = []
    for row in rows:
        item = RawOpportunityLineItem(
            id=uuid.uuid4(),
            snapshot_id=snapshot_id,
            sf_opportunity_id=row["sf_opportunity_id"],
            sf_line_item_id=row["sf_line_item_id"],
            opportunity_name=row["opportunity_name"],
            account_name=row["account_name"],
            opportunity_owner=row["opportunity_owner"],
            opportunity_type=row["opportunity_type"],
            channel_type=row["channel_type"],
            opportunity_amount=row["opportunity_amount"],
            close_date=row["close_date"],
            product_name=row["product_name"],
            product_code=row["product_code"],
            unit_price=row["unit_price"],
            quantity=row["quantity"],
            subscription_start_date=row["subscription_start_date"],
            subscription_end_date=row["subscription_end_date"],
            licence_period_months=row["licence_period_months"],
            business_line=row["business_line"],
        )
        session.add(item)
        inserted.append(item)
    session.flush()
    return inserted


def run_calculation_and_store(session, snapshot, raw_items):
    """Run ARR calculator and persist arr_line_items + arr_monthly_summary + alerts."""
    from app.backend.db.models import (
        ARRLineItem, ARRMonthlySummary, SnapshotAlert, ProductClassification, ConsultantCountry
    )
    from app.backend.core.arr_calculator import ARRCalculator, RawLineItem
    from decimal import Decimal

    # Load lookup tables from DB
    products = {p.product_name: p.product_type
                for p in session.query(ProductClassification).all()}
    countries = {c.consultant_name: c.country
                 for c in session.query(ConsultantCountry).all()}

    calc = ARRCalculator(products, countries)

    # Convert DB objects to domain objects
    domain_items = []
    raw_by_id = {}
    for raw in raw_items:
        domain = RawLineItem(
            sf_opportunity_id=str(raw.sf_opportunity_id),
            sf_line_item_id=str(raw.sf_line_item_id),
            opportunity_name=raw.opportunity_name or "",
            account_name=raw.account_name or "",
            opportunity_owner=raw.opportunity_owner or "",
            opportunity_type=raw.opportunity_type or "",
            channel_type=raw.channel_type or "",
            close_date=raw.close_date,
            product_name=raw.product_name,
            unit_price=Decimal(str(raw.unit_price)),
            quantity=Decimal(str(raw.quantity)),
            subscription_start_date=raw.subscription_start_date,
            subscription_end_date=raw.subscription_end_date,
            licence_period_months=raw.licence_period_months,
            business_line=raw.business_line,
            opportunity_amount=Decimal(str(raw.opportunity_amount)) if raw.opportunity_amount else None,
            product_code=raw.product_code,
        )
        domain_items.append(domain)
        raw_by_id[raw.sf_line_item_id] = raw.id

    arr_snapshot = calc.process_all(domain_items)

    # Persist arr_line_items
    for i, result in enumerate(arr_snapshot.line_items):
        raw_db_id = raw_by_id[result.raw.sf_line_item_id]
        ali = ARRLineItem(
            id=uuid.uuid4(),
            snapshot_id=snapshot.id,
            raw_line_item_id=raw_db_id,
            product_type=result.product_type,
            is_saas=result.is_saas,
            effective_start_date=result.effective_start_date,
            effective_end_date=result.effective_end_date,
            used_start_fallback=result.used_start_fallback,
            used_end_fallback=result.used_end_fallback,
            start_month=result.start_month,
            end_month_normalized=result.end_month_normalized,
            service_days=result.service_days,
            real_price=result.real_price,
            daily_price=result.daily_price,
            annualized_value=result.annualized_value,
            consultant_country=result.consultant_country,
            data_quality_flags=result.data_quality_flags,
        )
        session.add(ali)

    session.flush()
    print(f"  ARR line items stored: {len(arr_snapshot.line_items)}")

    # Build monthly summary
    all_months = _all_months_in_snapshot(arr_snapshot)
    monthly = calc.build_monthly_summary(arr_snapshot, all_months)
    for month_start, by_product in monthly.items():
        for product_type, arr_value in by_product.items():
            # Count active items for this product/month
            from datetime import timedelta
            month_end = month_start.replace(month=month_start.month % 12 + 1, day=1) - timedelta(days=1) \
                if month_start.month < 12 else month_start.replace(day=31)
            count = sum(
                1 for item in arr_snapshot.saas_items()
                if item.product_type == product_type
                and item.start_month <= month_end
                and item.end_month_normalized >= month_start
            )
            session.add(ARRMonthlySummary(
                snapshot_id=snapshot.id,
                month=month_start,
                product_type=product_type,
                arr_value=arr_value,
                line_items_count=count,
            ))
    session.flush()
    print(f"  Monthly summary rows: {sum(len(v) for v in monthly.values())}")

    # Persist alerts
    for alert in arr_snapshot.alerts:
        session.add(SnapshotAlert(
            id=uuid.uuid4(),
            snapshot_id=snapshot.id,
            alert_type=alert["alert_type"],
            severity=alert["severity"],
            sf_opportunity_id=alert.get("sf_opportunity_id"),
            opportunity_name=alert.get("opportunity_name"),
            account_name=alert.get("account_name"),
            product_name=alert.get("product_name"),
            description=alert["description"],
        ))
    session.flush()
    print(f"  Alerts stored: {len(arr_snapshot.alerts)}")

    # Update snapshot counts
    from app.backend.core.alert_checker import summarize_snapshot_quality
    quality = summarize_snapshot_quality(arr_snapshot)
    snapshot.unclassified_products_count = quality["unclassified_products"]
    snapshot.alerts_count = quality["total_alerts"]
    snapshot.sf_records_fetched = len(raw_items)
    snapshot.sf_records_processed = len(arr_snapshot.line_items)

    return arr_snapshot


def _all_months_in_snapshot(arr_snapshot) -> list:
    """Return sorted list of first-day-of-month dates covering all line items."""
    months = set()
    for item in arr_snapshot.saas_items():
        m = item.start_month
        while m <= item.end_month_normalized:
            months.add(m)
            # Advance one month
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)
    return sorted(months)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import Excel ARR data into the database")
    parser.add_argument("--excel", default=str(EXCEL_DEFAULT), help="Path to the Excel file")
    parser.add_argument("--env", default=".env", help="Path to .env file")
    args = parser.parse_args()

    # Load env
    from dotenv import load_dotenv
    load_dotenv(args.env)

    print(f"Loading Excel: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

    print("Reading lookup tables...")
    products = load_product_classifications(wb)
    countries = load_consultant_countries(wb)
    print(f"  Products found: {len(products)}, Consultants found: {len(countries)}")

    print("Reading opportunity line items...")
    rows = list(load_opos_rows(wb))
    print(f"  Rows read: {len(rows)}")

    from app.backend.db.connection import SessionLocal
    session = SessionLocal()
    t0 = datetime.now()

    try:
        print("Upserting lookup tables...")
        upsert_product_classifications(session, products)
        upsert_consultant_countries(session, countries)

        print("Creating snapshot...")
        snapshot = create_snapshot(session)
        print(f"  Snapshot ID: {snapshot.id}")

        print("Inserting raw line items...")
        raw_items = insert_raw_items(session, snapshot.id, rows)
        print(f"  Inserted: {len(raw_items)}")

        print("Running ARR calculation...")
        arr_snapshot = run_calculation_and_store(session, snapshot, raw_items)

        duration = (datetime.now() - t0).total_seconds()
        snapshot.status = "success"
        snapshot.duration_seconds = duration

        session.commit()
        print(f"\nDone in {duration:.1f}s. Snapshot {snapshot.id} created successfully.")

    except Exception as exc:
        session.rollback()
        print(f"\nERROR: {exc}")
        raise
    finally:
        session.close()


if __name__ == "__main__":
    main()
