"""
API endpoint tests for Phase B.

Uses FastAPI TestClient with a SQLite in-memory database so tests
run without Docker/PostgreSQL. The get_db dependency is overridden.
"""

import sys
import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.backend.db.connection import Base, get_db
from app.backend.db.models import (
    ARRLineItem,
    ARRMonthlySummary,
    ConsultantCountry,
    ProductClassification,
    RawOpportunityLineItem,
    Snapshot,
    SnapshotAlert,
)
from app.backend.core.arr_calculator import RawLineItem
from app.backend.main import app

# ---------------------------------------------------------------------------
# Test database setup (SQLite in-memory)
# ---------------------------------------------------------------------------

SQLALCHEMY_TEST_URL = "sqlite://"

engine = create_engine(
    SQLALCHEMY_TEST_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)

# Enable JSON support for SQLite (data_quality_flags column)
import json
from sqlalchemy import TypeDecorator, Text as SAText


class JSONBCompat(TypeDecorator):
    """Stores JSONB as JSON text in SQLite."""
    impl = SAText
    cache_ok = True

    def process_bind_param(self, value, dialect):
        if value is None:
            return "[]"
        return json.dumps(value)

    def process_result_value(self, value, dialect):
        if value is None:
            return []
        return json.loads(value)


# Patch JSONB columns to use JSONBCompat in tests
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy import event as sa_event


@sa_event.listens_for(engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    cursor = dbapi_connection.cursor()
    cursor.execute("PRAGMA journal_mode=WAL")
    cursor.close()


TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db


@pytest.fixture(autouse=True)
def setup_db():
    """Create tables before each test and drop after."""
    # SQLite doesn't support JSONB natively — patch columns in models
    for table in Base.metadata.tables.values():
        for col in table.columns:
            if hasattr(col.type, "__class__") and col.type.__class__.__name__ == "JSONB":
                col.type = JSONBCompat()
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


@pytest.fixture
def client():
    return TestClient(app)


# ---------------------------------------------------------------------------
# Helpers to seed test data
# ---------------------------------------------------------------------------

def _make_snapshot(db, status="completed", sync_type="excel_import"):
    snap = Snapshot(
        id=uuid.uuid4(),
        created_at=datetime(2026, 4, 17),
        sync_type=sync_type,
        status=status,
        sf_records_fetched=2,
        sf_records_processed=2,
        alerts_count=1,
        unclassified_products_count=0,
    )
    db.add(snap)
    db.flush()
    return snap


def _make_raw(db, snap_id, sf_opp_id="OPP001", sf_line_id="LI001", product="SaaS LMS"):
    row_id = uuid.uuid4()
    raw = RawOpportunityLineItem(
        id=row_id,
        snapshot_id=snap_id,
        sf_opportunity_id=sf_opp_id,
        sf_line_item_id=sf_line_id,
        opportunity_name="Test Opp",
        account_name="Acme Corp",
        opportunity_owner="Miguel V.",
        opportunity_type="nuevo_negocio",
        channel_type="KAM",
        close_date=date(2025, 1, 15),
        product_name=product,
        unit_price=Decimal("10000.00"),
        quantity=Decimal("1"),
        subscription_start_date=date(2025, 1, 1),
        subscription_end_date=date(2025, 12, 31),
    )
    db.add(raw)
    db.flush()
    return raw


def _make_arr(
    db,
    snap_id,
    raw_id,
    product_type="SaaS LMS",
    arr=Decimal("10000"),
    start_month=date(2025, 1, 1),
    end_month_normalized=date(2025, 12, 31),
    excluded_from_arr=False,
):
    arr_item = ARRLineItem(
        id=uuid.uuid4(),
        snapshot_id=snap_id,
        raw_line_item_id=raw_id,
        product_type=product_type,
        is_saas=True,
        effective_start_date=start_month,
        effective_end_date=end_month_normalized,
        used_start_fallback=False,
        used_end_fallback=False,
        start_month=start_month,
        end_month_normalized=end_month_normalized,
        service_days=(end_month_normalized - start_month).days or 1,
        real_price=Decimal("10000"),
        daily_price=Decimal("27.47"),
        annualized_value=arr,
        consultant_country="Spain",
        data_quality_flags=[],
        excluded_from_arr=excluded_from_arr,
    )
    db.add(arr_item)
    db.flush()
    return arr_item


def _make_summary(db, snap_id, month, product_type="SaaS LMS", arr=Decimal("10000")):
    row = ARRMonthlySummary(
        snapshot_id=snap_id,
        month=month,
        product_type=product_type,
        arr_value=arr,
        line_items_count=1,
    )
    db.add(row)
    db.flush()
    return row


def _build_excel_bytes() -> bytes:
    wb = openpyxl.Workbook()

    ws_products = wb.active
    ws_products.title = "Productos SF SAAS"
    ws_products.append(
        [
            "A",
            "Nombre producto",
            "Codigo",
            "D",
            "Linea negocio",
            "Categoria",
            "Subcategoria",
            "Tipo producto",
        ]
    )
    ws_products.append(
        ["", "Licencias LMS", "LMS-001", "", "isEazy LMS", "SaaS", "LMS", "SaaS LMS"]
    )

    ws_consultants = wb.create_sheet("País Consultor")
    ws_consultants.append(["", "", "", ""])
    ws_consultants.append(["", "", "", ""])
    ws_consultants.append(["", "", "", ""])
    ws_consultants.append(["", "", "Consultor", "Pais"])
    ws_consultants.append(["", "", "Maria Lopez", "Spain"])

    ws_opos = wb.create_sheet("Opos con Productos")
    ws_opos.append(
        [
            "Propietario",
            "Cuenta",
            "Oportunidad",
            "Tipo",
            "Canal",
            "Importe",
            "Fecha cierre",
            "Creacion",
            "Etapa",
            "Producto",
            "Precio",
            "Inicio",
            "Fin",
            "Meses",
            "Linea negocio",
            "Cantidad",
            "Codigo",
            "Creado por",
        ]
    )
    ws_opos.append(
        [
            "Maria Lopez",
            "Acme Corp",
            "Expansion ACME",
            "New Business",
            "Inbound",
            12000,
            "15/01/2026",
            "10/01/2026",
            "Ganada",
            "Licencias LMS",
            12000,
            "01/01/2026",
            "31/12/2026",
            12,
            "isEazy LMS",
            1,
            "LMS-001",
            "Maria",
        ]
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_raw_salesforce_excel_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opos con Productos"
    ws.append(
        [
            "Propietario de oportunidad",
            "Nombre de la cuenta",
            "Nombre de la oportunidad",
            "Tipo",
            "Tipo de oportunidad",
            "Importe",
            "Fecha de cierre",
            "Fecha de creación",
            "Etapa",
            "Nombre del producto",
            "Precio de venta",
            "Subscription Start Date",
            "Subscription End Date",
            "Licence period (months)",
            "Línea de negocio",
            "Cantidad",
            "Product",
            "Creado por",
        ]
    )
    ws.append(
        [
            "Maria Lopez",
            "Acme Corp",
            "Renewal ACME",
            "Negocio existente",
            "Inbound",
            12000,
            "15/01/2026",
            "10/01/2026",
            "Ganada",
            "Usuarios LMS",
            12000,
            "01/01/2026",
            "31/12/2026",
            12,
            "isEazy LMS",
            1,
            "LMS-RAW",
            "Maria Lopez",
        ]
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

def test_health(client):
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


# ---------------------------------------------------------------------------
# Snapshots
# ---------------------------------------------------------------------------

def test_list_snapshots_empty(client):
    r = client.get("/api/snapshots")
    assert r.status_code == 200
    assert r.json() == []


def test_list_snapshots_returns_data(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    db.commit()
    db.close()

    r = client.get("/api/snapshots")
    assert r.status_code == 200
    data = r.json()
    assert len(data) == 1
    assert data[0]["status"] == "completed"


def test_get_snapshot_not_found(client):
    r = client.get(f"/api/snapshots/{uuid.uuid4()}")
    assert r.status_code == 404


def test_get_snapshot_found(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    db.commit()
    snap_id = snap.id
    db.close()

    r = client.get(f"/api/snapshots/{snap_id}")
    assert r.status_code == 200
    assert r.json()["id"] == str(snap_id)


# ---------------------------------------------------------------------------
# Sync
# ---------------------------------------------------------------------------

def test_sync_returns_500_when_salesforce_config_is_missing(client):
    r = client.post("/api/sync", json={"triggered_by": "test"})
    assert r.status_code == 500
    assert "Missing Salesforce configuration" in r.json()["detail"]


def test_sync_creates_snapshot_from_salesforce_extractor(client, monkeypatch):
    db = TestingSessionLocal()
    db.add(ProductClassification(product_name="Licencias LMS", product_type="SaaS LMS"))
    db.add(ConsultantCountry(consultant_name="Maria Lopez", country="Spain"))
    db.commit()
    db.close()

    raw_items = [
        RawLineItem(
            sf_opportunity_id="006OPP001234567",
            sf_line_item_id="00kLINEITEM001",
            opportunity_name="Expansion ACME",
            account_name="ACME Corp",
            opportunity_owner="Maria Lopez",
            opportunity_type="New Business",
            channel_type="Inbound",
            close_date=date(2026, 1, 15),
            product_name="Licencias LMS",
            unit_price=Decimal("12000"),
            quantity=Decimal("1"),
            subscription_start_date=date(2026, 1, 1),
            subscription_end_date=date(2026, 12, 31),
            licence_period_months=12,
            business_line="isEazy LMS",
            opportunity_amount=Decimal("12000"),
            product_code="LMS-001",
        )
    ]

    from app.backend.api.routes import sync as sync_route

    class DummyExtractor:
        def fetch_raw_line_items(self):
            return raw_items

    monkeypatch.setattr(sync_route, "SalesforceExtractor", lambda: DummyExtractor())

    r = client.post("/api/sync", json={"triggered_by": "test"})
    assert r.status_code == 200
    data = r.json()
    assert data["status"] == "completed"
    assert data["records_processed"] == 1


def test_import_excel_rejects_non_xlsx(client):
    r = client.post(
        "/api/imports/excel",
        files={"file": ("arr.csv", b"fake", "text/csv")},
    )
    assert r.status_code == 400
    assert "Solo se admiten ficheros .xlsx" in r.json()["detail"]


def test_import_excel_creates_completed_snapshot(client):
    excel_bytes = _build_excel_bytes()

    r = client.post(
        "/api/imports/excel",
        files={
            "file": (
                "ARR Oportunidad.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert r.status_code == 200
    data = r.json()
    assert data["status"] == "completed"
    assert data["records_processed"] == 1

    snapshots = client.get("/api/snapshots")
    assert snapshots.status_code == 200
    assert snapshots.json()[0]["sync_type"] == "excel_import"

    summary = client.get("/api/arr/summary")
    assert summary.status_code == 200
    assert len(summary.json()["months"]) >= 1


def test_import_raw_salesforce_export_without_calculated_sheets(client):
    excel_bytes = _build_raw_salesforce_excel_bytes()

    r = client.post(
        "/api/imports/excel",
        files={
            "file": (
                "Opos con Productos.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert r.status_code == 200
    data = r.json()
    assert data["status"] == "completed"
    assert data["records_processed"] == 1

    summary = client.get("/api/arr/summary")
    assert summary.status_code == 200
    months = summary.json()["months"]
    assert len(months) >= 1
    assert Decimal(months[-1]["by_product_type"]["SaaS LMS"]) > 0


# ---------------------------------------------------------------------------
# ARR Summary
# ---------------------------------------------------------------------------

def test_arr_summary_no_snapshot(client):
    r = client.get("/api/arr/summary")
    assert r.status_code == 404


def test_arr_summary_returns_months(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    # Two non-overlapping contracts: Jan-only and Feb-only
    raw_jan = _make_raw(db, snap.id, sf_opp_id="OPP_JAN", sf_line_id="LI_JAN")
    raw_jan.subscription_start_date = date(2025, 1, 1)
    raw_jan.subscription_end_date = date(2025, 1, 31)
    raw_feb = _make_raw(db, snap.id, sf_opp_id="OPP_FEB", sf_line_id="LI_FEB")
    raw_feb.subscription_start_date = date(2025, 2, 1)
    raw_feb.subscription_end_date = date(2025, 2, 28)
    db.flush()
    _make_arr(
        db, snap.id, raw_jan.id, arr=Decimal("12000"),
        start_month=date(2025, 1, 1), end_month_normalized=date(2025, 1, 30),
    )
    _make_arr(
        db, snap.id, raw_feb.id, arr=Decimal("13000"),
        start_month=date(2025, 2, 1), end_month_normalized=date(2025, 2, 27),
    )
    db.commit()
    db.close()

    r = client.get("/api/arr/summary")
    assert r.status_code == 200
    data = r.json()
    assert len(data["months"]) == 2
    assert Decimal(data["months"][0]["total_arr"]) == Decimal("12000")
    assert Decimal(data["months"][1]["mom_change"]) == Decimal("1000")


def test_arr_summary_filter_by_product_type(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    raw_lms = _make_raw(db, snap.id, sf_opp_id="OPP_LMS", sf_line_id="LI_LMS")
    raw_skills = _make_raw(db, snap.id, sf_opp_id="OPP_SKL", sf_line_id="LI_SKL")
    db.flush()
    # Both items span only Jan 2025 to keep the test focused on the filter
    _make_arr(
        db, snap.id, raw_lms.id, product_type="SaaS LMS", arr=Decimal("12000"),
        start_month=date(2025, 1, 1), end_month_normalized=date(2025, 1, 30),
    )
    _make_arr(
        db, snap.id, raw_skills.id, product_type="SaaS Skills", arr=Decimal("8000"),
        start_month=date(2025, 1, 1), end_month_normalized=date(2025, 1, 30),
    )
    db.commit()
    db.close()

    r = client.get("/api/arr/summary?product_type=SaaS+LMS")
    assert r.status_code == 200
    data = r.json()
    assert len(data["months"]) == 1
    assert Decimal(data["months"][0]["total_arr"]) == Decimal("12000")


# ---------------------------------------------------------------------------
# ARR Line Items
# ---------------------------------------------------------------------------

def test_arr_line_items_empty(client):
    r = client.get("/api/arr/line-items")
    assert r.status_code == 404


def test_arr_line_items_pagination(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    for i in range(5):
        raw = _make_raw(db, snap.id, sf_opp_id=f"OPP{i:03d}", sf_line_id=f"LI{i:03d}")
        _make_arr(db, snap.id, raw.id)
    db.commit()
    db.close()

    r = client.get("/api/arr/line-items?page=1&page_size=3")
    assert r.status_code == 200
    data = r.json()
    assert data["total"] == 5
    assert len(data["items"]) == 3
    assert data["page"] == 1


# ---------------------------------------------------------------------------
# Alerts
# ---------------------------------------------------------------------------

def test_alerts_empty(client):
    r = client.get("/api/alerts")
    assert r.status_code == 200
    assert r.json() == []


def test_alerts_list_and_filter(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    alert = SnapshotAlert(
        id=uuid.uuid4(),
        snapshot_id=snap.id,
        alert_type="UNCLASSIFIED_PRODUCT",
        severity="error",
        description="Producto no clasificado",
        reviewed=False,
    )
    other_alert = SnapshotAlert(
        id=uuid.uuid4(),
        snapshot_id=snap.id,
        alert_type="MISSING_END_DATE",
        severity="warning",
        description="Falta fecha fin",
        reviewed=True,
    )
    db.add(alert)
    db.add(other_alert)
    db.commit()
    db.close()

    r = client.get("/api/alerts?reviewed=false")
    assert r.status_code == 200
    assert len(r.json()) == 1

    r2 = client.get("/api/alerts?reviewed=true")
    assert r2.status_code == 200
    assert len(r2.json()) == 1

    r3 = client.get("/api/alerts?alert_type=UNCLASSIFIED_PRODUCT")
    assert r3.status_code == 200
    assert len(r3.json()) == 1
    assert r3.json()[0]["alert_type"] == "UNCLASSIFIED_PRODUCT"


def test_patch_alert(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    alert = SnapshotAlert(
        id=uuid.uuid4(),
        snapshot_id=snap.id,
        alert_type="UNCLASSIFIED_PRODUCT",
        severity="error",
        description="Producto no clasificado",
        reviewed=False,
    )
    db.add(alert)
    db.commit()
    alert_id = alert.id
    db.close()

    r = client.patch(
        f"/api/alerts/{alert_id}",
        json={"reviewed": True, "review_note": "Ok, producto nuevo", "reviewed_by": "dmunoz"},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["reviewed"] is True
    assert data["review_note"] == "Ok, producto nuevo"
    assert data["reviewed_by"] == "dmunoz"
    assert data["reviewed_at"] is not None


# ---------------------------------------------------------------------------
# Config: Products
# ---------------------------------------------------------------------------

def test_list_products_empty(client):
    r = client.get("/api/config/products")
    assert r.status_code == 200
    assert r.json() == []


def test_create_and_list_product(client):
    payload = {
        "product_name": "Usuarios LMS",
        "product_type": "SaaS LMS",
        "business_line": "isEazy LMS",
    }
    r = client.post("/api/config/products", json=payload)
    assert r.status_code == 201
    data = r.json()
    assert data["product_name"] == "Usuarios LMS"
    assert data["is_saas"] is True

    r2 = client.get("/api/config/products")
    assert len(r2.json()) == 1


def test_create_product_duplicate(client):
    payload = {"product_name": "Usuarios LMS", "product_type": "SaaS LMS"}
    client.post("/api/config/products", json=payload)
    r = client.post("/api/config/products", json=payload)
    assert r.status_code == 409


def test_update_product(client):
    r = client.post("/api/config/products", json={"product_name": "Old", "product_type": "SaaS LMS"})
    product_id = r.json()["id"]

    r2 = client.put(f"/api/config/products/{product_id}", json={"product_type": "Implantación"})
    assert r2.status_code == 200
    assert r2.json()["product_type"] == "Implantación"
    assert r2.json()["is_saas"] is False


# ---------------------------------------------------------------------------
# Config: Consultants
# ---------------------------------------------------------------------------

def test_list_consultants_empty(client):
    r = client.get("/api/config/consultants")
    assert r.status_code == 200
    assert r.json() == []


def test_update_consultant_not_found(client):
    r = client.put("/api/config/consultants/999", json={"country": "France"})
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# Stripe MRR
# ---------------------------------------------------------------------------

def test_stripe_mrr_empty(client):
    r = client.get("/api/stripe-mrr")
    assert r.status_code == 200
    assert r.json() == []


def test_stripe_mrr_upsert(client):
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    db.commit()
    snap_id = snap.id
    db.close()

    payload = {
        "snapshot_id": str(snap_id),
        "month": "2026-01-01",
        "mrr": "9200.00",
        "entered_by": "dmunoz",
    }
    r = client.put("/api/stripe-mrr", json=payload)
    assert r.status_code == 200
    data = r.json()
    assert data["mrr"] == "9200.00"
    assert data["arr_equivalent"] == "110400.00"

    # Upsert again — should update
    payload["mrr"] = "9500.00"
    r2 = client.put("/api/stripe-mrr", json=payload)
    assert r2.status_code == 200
    assert r2.json()["mrr"] == "9500.00"

    # GET returns the entry
    r3 = client.get(f"/api/stripe-mrr?snapshot_id={snap_id}")
    assert len(r3.json()) == 1
    assert r3.json()[0]["mrr"] == "9500.00"


def test_cron_daily_sync_rejects_wrong_secret(client, monkeypatch):
    monkeypatch.setenv("CRON_SECRET", "correct-secret")
    r = client.post("/api/sync/cron/daily", headers={"x-cron-secret": "wrong"})
    assert r.status_code == 401


# ---------------------------------------------------------------------------
# ARR mode=from_close
# ---------------------------------------------------------------------------

def test_arr_summary_from_close_mode(client):
    """mode=from_close uses close_date as the start, potentially shifting months."""
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    # Contract: service starts Jan 2025, but close_date was Nov 2024
    raw = _make_raw(db, snap.id, sf_opp_id="OPP_CLOSE", sf_line_id="LI_CLOSE")
    raw.close_date = date(2024, 11, 1)
    raw.subscription_start_date = date(2025, 1, 1)
    raw.subscription_end_date = date(2025, 12, 31)
    db.flush()
    _make_arr(
        db, snap.id, raw.id,
        arr=Decimal("10000"),
        start_month=date(2025, 1, 1),
        end_month_normalized=date(2025, 12, 30),
    )
    db.commit()
    db.close()

    r_start = client.get("/api/arr/summary?mode=from_start")
    assert r_start.status_code == 200
    first_month_start = r_start.json()["months"][0]["month"]

    r_close = client.get("/api/arr/summary?mode=from_close")
    assert r_close.status_code == 200
    first_month_close = r_close.json()["months"][0]["month"]

    # from_close should start in Nov 2024, from_start in Jan 2025
    assert first_month_close < first_month_start


# ---------------------------------------------------------------------------
# ARR line item exclusion
# ---------------------------------------------------------------------------

def test_patch_line_item_exclude(client):
    """PATCH /api/arr/line-items/{id} toggles excluded_from_arr."""
    db = TestingSessionLocal()
    snap = _make_snapshot(db)
    raw = _make_raw(db, snap.id)
    arr_item = _make_arr(db, snap.id, raw.id, arr=Decimal("10000"))
    db.commit()
    arr_id = arr_item.id
    db.close()

    # Initially included — summary returns data
    r1 = client.get("/api/arr/summary")
    assert r1.status_code == 200
    assert len(r1.json()["months"]) > 0

    # Exclude the item
    r_patch = client.patch(f"/api/arr/line-items/{arr_id}", json={"excluded_from_arr": True})
    assert r_patch.status_code == 200
    assert r_patch.json()["excluded_from_arr"] is True

    # Summary should now be empty
    r2 = client.get("/api/arr/summary")
    assert r2.status_code == 200
    assert len(r2.json()["months"]) == 0

    # Re-include
    r_patch2 = client.patch(f"/api/arr/line-items/{arr_id}", json={"excluded_from_arr": False})
    assert r_patch2.json()["excluded_from_arr"] is False

    r3 = client.get("/api/arr/summary")
    assert len(r3.json()["months"]) > 0


def test_patch_line_item_not_found(client):
    r = client.patch(f"/api/arr/line-items/{uuid.uuid4()}", json={"excluded_from_arr": True})
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# Overlap alerts generated on sync
# ---------------------------------------------------------------------------

def test_sync_generates_overlap_alerts(client, monkeypatch):
    """Two overlapping contracts for same account+product generate OVERLAPPING_CONTRACTS alerts."""
    db = TestingSessionLocal()
    db.add(ProductClassification(product_name="Licencias LMS", product_type="SaaS LMS"))
    db.add(ConsultantCountry(consultant_name="Maria Lopez", country="Spain"))
    db.commit()
    db.close()

    from app.backend.core.arr_calculator import RawLineItem as RL
    raw_items = [
        RL(
            sf_opportunity_id="OPP_A",
            sf_line_item_id="LI_A",
            opportunity_name="Contrato A",
            account_name="Acme Corp",
            opportunity_owner="Maria Lopez",
            opportunity_type="New Business",
            channel_type="Inbound",
            close_date=date(2025, 1, 1),
            product_name="Licencias LMS",
            unit_price=Decimal("12000"),
            quantity=Decimal("1"),
            subscription_start_date=date(2025, 1, 1),
            subscription_end_date=date(2025, 12, 31),
            licence_period_months=12,
            business_line="isEazy LMS",
        ),
        RL(
            sf_opportunity_id="OPP_B",
            sf_line_item_id="LI_B",
            opportunity_name="Contrato B",
            account_name="Acme Corp",
            opportunity_owner="Maria Lopez",
            opportunity_type="Renewal",
            channel_type="Inbound",
            close_date=date(2025, 6, 1),
            product_name="Licencias LMS",
            unit_price=Decimal("13000"),
            quantity=Decimal("1"),
            subscription_start_date=date(2025, 6, 1),
            subscription_end_date=date(2026, 5, 31),
            licence_period_months=12,
            business_line="isEazy LMS",
        ),
    ]

    from app.backend.api.routes import sync as sync_route

    class DummyExtractor:
        def fetch_raw_line_items(self):
            return raw_items

    monkeypatch.setattr(sync_route, "SalesforceExtractor", lambda: DummyExtractor())

    r = client.post("/api/sync", json={"triggered_by": "test"})
    assert r.status_code == 200

    alerts_r = client.get("/api/alerts?alert_type=OVERLAPPING_CONTRACTS")
    assert alerts_r.status_code == 200
    overlap_alerts = alerts_r.json()
    assert len(overlap_alerts) == 2
    for alert in overlap_alerts:
        assert alert["arr_line_item_id"] is not None


def test_cron_daily_sync_skips_when_data_unchanged(client, monkeypatch):
    monkeypatch.setenv("CRON_SECRET", "correct-secret")

    db = TestingSessionLocal()
    db.add(ProductClassification(product_name="Licencias LMS", product_type="SaaS LMS"))
    db.add(ConsultantCountry(consultant_name="Maria Lopez", country="Spain"))
    db.commit()
    db.close()

    raw_items = [
        RawLineItem(
            sf_opportunity_id="006OPP999",
            sf_line_item_id="00kLINE999",
            opportunity_name="Renewal ACME",
            account_name="ACME Corp",
            opportunity_owner="Maria Lopez",
            opportunity_type="Renewal",
            channel_type="Inbound",
            close_date=date(2026, 1, 15),
            product_name="Licencias LMS",
            unit_price=Decimal("12000"),
            quantity=Decimal("1"),
            subscription_start_date=date(2026, 1, 1),
            subscription_end_date=date(2026, 12, 31),
            licence_period_months=12,
            business_line="isEazy LMS",
            opportunity_amount=Decimal("12000"),
            product_code="LMS-001",
        )
    ]

    from app.backend.api.routes import sync as sync_route
    from app.backend.core.snapshot_manager import compute_raw_hash

    class DummyExtractor:
        def fetch_raw_line_items(self):
            return raw_items

    monkeypatch.setattr(sync_route, "SalesforceExtractor", lambda: DummyExtractor())

    # First sync creates snapshot
    r1 = client.post("/api/sync/cron/daily", headers={"x-cron-secret": "correct-secret"})
    assert r1.status_code == 200
    assert r1.json()["status"] == "completed"
    assert r1.json()["skipped"] is False

    # Second sync with identical data must be skipped
    r2 = client.post("/api/sync/cron/daily", headers={"x-cron-secret": "correct-secret"})
    assert r2.status_code == 200
    assert r2.json()["status"] == "skipped"
    assert r2.json()["skipped"] is True
