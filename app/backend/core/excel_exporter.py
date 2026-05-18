"""Build a pivot-friendly xlsx export for a calculated ARR snapshot."""

import io
import re
from datetime import date, timedelta
from decimal import Decimal
from uuid import UUID

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.backend.db.models import ARRLineItem, RawOpportunityLineItem, Snapshot, SnapshotStripeMRR


_HEADER_FILL = PatternFill(start_color="2F185F", end_color="2F185F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATE_FORMAT = "yyyy-mm-dd"
_MONTH_FORMAT = "yyyy-mm"
_MONEY_FORMAT = '#,##0.00 "EUR"'
_ILLEGAL_EXCEL_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _write_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 42)


def _set_export_column_widths(ws) -> None:
    widths = {
        "A": 20,
        "B": 13,
        "C": 12,
        "D": 28,
        "E": 24,
        "F": 16,
        "G": 20,
        "H": 32,
        "I": 36,
        "J": 18,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 14,
        "Q": 14,
        "R": 12,
        "S": 20,
        "T": 20,
        "U": 38,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _cell_value(value):
    if isinstance(value, str):
        return _ILLEGAL_EXCEL_CHARS.sub("", value)
    return value


def _append_row(ws, values: list) -> None:
    ws.append([_cell_value(value) for value in values])


def _last_day_of_month(first_day: date) -> date:
    if first_day.month == 12:
        return first_day.replace(day=31)
    return first_day.replace(month=first_day.month + 1) - timedelta(days=1)


def _next_month(first_day: date) -> date:
    if first_day.month == 12:
        return first_day.replace(year=first_day.year + 1, month=1)
    return first_day.replace(month=first_day.month + 1)


def _month_range(start: date, end: date) -> list[date]:
    months: list[date] = []
    current = start.replace(day=1)
    end_first = end.replace(day=1)
    while current <= end_first:
        months.append(current)
        current = _next_month(current)
    return months


def _active_start_month(arr: ARRLineItem, raw: RawOpportunityLineItem, mode: str) -> date:
    if mode == "from_close":
        opp_type = (raw.opportunity_type or "").lower().strip()
        if (
            opp_type == "nuevo negocio"
            and raw.subscription_start_date
            and raw.close_date < raw.subscription_start_date
        ):
            return raw.close_date.replace(day=1)
    return arr.start_month


def _apply_formats(ws) -> None:
    date_headers = {"mes_calculo", "fecha_desde_arr", "fecha_hasta_arr", "fecha_inicio_bd", "fecha_fin_bd", "fecha_cierre"}
    money_headers = {"arr_eur", "precio_real"}
    headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}
    for header in date_headers:
        col = headers.get(header)
        if col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = _DATE_FORMAT if header != "mes_calculo" else _MONTH_FORMAT
    for header in money_headers:
        col = headers.get(header)
        if col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = _MONEY_FORMAT


def build_snapshot_excel(snapshot_id: UUID, db: Session) -> bytes:
    snap = db.query(Snapshot).filter(Snapshot.id == snapshot_id).first()
    if not snap:
        raise ValueError(f"Snapshot {snapshot_id} not found")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_calculated_arr(wb, snapshot_id, db, mode="from_start", sheet_name="Desde inicio")
    _sheet_calculated_arr(wb, snapshot_id, db, mode="from_close", sheet_name="Desde cierre")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _sheet_calculated_arr(wb, snapshot_id: UUID, db: Session, mode: str, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = [
        "metodologia",
        "mes_calculo",
        "source",
        "cliente",
        "consultor",
        "pais_consultor",
        "linea_negocio",
        "producto",
        "oportunidad",
        "tipo_oportunidad",
        "fecha_desde_arr",
        "fecha_hasta_arr",
        "fecha_inicio_bd",
        "fecha_fin_bd",
        "fecha_cierre",
        "arr_eur",
        "precio_real",
        "service_days",
        "sf_opportunity_id",
        "sf_line_item_id",
        "arr_line_item_id",
    ]
    _write_headers(ws, headers)

    rows = (
        db.query(ARRLineItem, RawOpportunityLineItem)
        .join(RawOpportunityLineItem, ARRLineItem.raw_line_item_id == RawOpportunityLineItem.id)
        .filter(
            ARRLineItem.snapshot_id == snapshot_id,
            ARRLineItem.is_saas == True,
            ARRLineItem.excluded_from_arr == False,
        )
        .order_by(RawOpportunityLineItem.account_name, ARRLineItem.start_month)
        .all()
    )

    method_label = "Desde fecha inicio" if mode == "from_start" else "Desde fecha cierre"
    for arr, raw in rows:
        active_start = _active_start_month(arr, raw, mode)
        active_end = arr.end_month_normalized
        if active_start > active_end:
            continue
        for month in _month_range(active_start, active_end):
            _append_row(ws, [
                method_label,
                month,
                "Salesforce",
                raw.account_name or "Sin cuenta",
                raw.opportunity_owner or "Sin consultor",
                arr.consultant_country or "Sin pais",
                arr.product_type or "Unknown",
                raw.product_name,
                raw.opportunity_name,
                raw.opportunity_type,
                active_start,
                active_end,
                arr.effective_start_date,
                arr.effective_end_date,
                raw.close_date,
                float(Decimal(str(arr.annualized_value or 0))),
                float(Decimal(str(arr.real_price or 0))),
                arr.service_days,
                raw.sf_opportunity_id,
                raw.sf_line_item_id,
                str(arr.id),
            ])

    stripe_rows = (
        db.query(SnapshotStripeMRR)
        .filter(SnapshotStripeMRR.snapshot_id == snapshot_id)
        .order_by(SnapshotStripeMRR.month)
        .all()
    )
    for stripe in stripe_rows:
        month = stripe.month.replace(day=1)
        _append_row(ws, [
            method_label,
            month,
            "Stripe",
            "Author Online Stripe",
            "[Author Online Stripe]",
            "-",
            "Author Online",
            "Author Online Stripe",
            "Author Online Stripe",
            None,
            month,
            month,
            month,
            _last_day_of_month(month),
            None,
            float(Decimal(str(stripe.mrr or 0))),
            float(Decimal(str(stripe.mrr or 0))),
            None,
            None,
            None,
            f"stripe-{month.isoformat()}",
        ])

    ws.auto_filter.ref = ws.dimensions
    _set_export_column_widths(ws)
