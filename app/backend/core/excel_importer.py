"""Shared Excel import pipeline for manual ARR snapshot creation."""

from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from app.backend.core.alert_checker import summarize_snapshot_quality
from app.backend.core.arr_calculator import ARRCalculator, RawLineItem
from app.backend.db.models import (
    ARRLineItem,
    ARRMonthlySummary,
    ConsultantCountry,
    ProductClassification,
    RawOpportunityLineItem,
    Snapshot,
    SnapshotAlert,
)


class ExcelImportError(Exception):
    """Raised when a workbook cannot be parsed or imported."""


@dataclass
class ExcelImportSummary:
    snapshot: Snapshot
    rows_read: int


BUSINESS_LINE_TO_PRODUCT_TYPE = {
    "iseazy lms": "SaaS LMS",
    "lms": "SaaS LMS",
    "iseazy author": "SaaS Author",
    "author": "SaaS Author",
    "iseazy skills": "SaaS Skills",
    "skills": "SaaS Skills",
    "iseazy engage": "SaaS Engage",
    "engage": "SaaS Engage",
    "iseazy aio": "SaaS AIO",
    "aio": "SaaS AIO",
    "author online": "Author Online",
}

UNASSIGNED_MASTER_VALUE = "[SIN ASIGNAR]"

RECURRING_PRODUCT_KEYWORDS = (
    "licencia",
    "licencias",
    "usuario",
    "usuarios",
    "subscription",
    "suscripcion",
    "suscripción",
    "saas",
    "plataforma",
)


def _parse_date_text(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        n = float(s)
        return date(1899, 12, 30) + timedelta(days=int(n))
    except (ValueError, TypeError):
        return None


def _decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def _str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_key(value: str | None) -> str:
    if value is None:
        return ""
    normalized = value.strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ñ": "n",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def _row_header_map(ws) -> dict[str, int]:
    try:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return {}
    return {_normalize_key(_str(value)): index for index, value in enumerate(header)}


def _cell(row, headers: dict[str, int], *names: str):
    for name in names:
        index = headers.get(_normalize_key(name))
        if index is not None and index < len(row):
            return row[index]
    return None


def _looks_recurring_product(row: dict) -> bool:
    if row.get("subscription_start_date") or row.get("subscription_end_date"):
        return True
    if row.get("licence_period_months"):
        return True
    product_name = _normalize_key(row.get("product_name"))
    return any(keyword in product_name for keyword in RECURRING_PRODUCT_KEYWORDS)


def _infer_product_type(product_name: str, business_line: str | None, row: dict) -> Optional[str]:
    combined = f"{_normalize_key(business_line)} {_normalize_key(product_name)}"
    for key, product_type in BUSINESS_LINE_TO_PRODUCT_TYPE.items():
        if _normalize_key(key) in combined and _looks_recurring_product(row):
            return product_type
    return None


def load_product_classifications(wb) -> dict:
    """
    Returns {product_name: classification_info}.

    When the same product name appears in multiple business lines, the lookup
    key is enriched internally using the business_line (col 4 = Línea de Negocio2,
    which uses the same format as the BBDD, e.g. 'isEazy LMS'). The returned dict
    uses compound keys "(name)|(business_line)" so callers can look up by both.
    A plain product_name key is also kept as fallback (first occurrence wins).
    """
    if "Productos SF SAAS" not in wb.sheetnames:
        return {}

    ws = wb["Productos SF SAAS"]

    result = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or len(row) < 2 or row[1] is None:
            continue
        name = _str(row[1])
        if not name:
            continue
        business_line = _str(row[4]) if len(row) > 4 else None
        info = {
            "product_code": _str(row[2]) if len(row) > 2 else None,
            "business_line": business_line,
            "category": _str(row[5]) if len(row) > 5 else None,
            "subcategory": _str(row[6]) if len(row) > 6 else None,
            "product_type": _str(row[7]) if len(row) > 7 else None,
        }
        # Compound key for exact (name, business_line) lookups
        if business_line:
            compound = f"{name}|{business_line}"
            result[compound] = info
        # Plain name fallback — only first occurrence, so we don't overwrite with a
        # different classification for the same product in another business line
        if name not in result:
            result[name] = info
    return result


def load_consultant_countries(wb) -> dict:
    sheet_names = ["PaÃ­s Consultor", "País Consultor"]
    ws = None
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    if ws is None:
        return {}

    result = {}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[2] == "Consultor":
                header_found = True
            continue
        name = _str(row[2])
        country = _str(row[3])
        if name and country:
            result[name] = country
    return result


def load_opos_rows(wb):
    try:
        ws = wb["Opos con Productos"]
    except KeyError as exc:
        raise ExcelImportError("La hoja 'Opos con Productos' no existe en el Excel.") from exc

    headers = _row_header_map(ws)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue

        if headers:
            product_name = _str(
                _cell(row, headers, "Nombre del producto", "Producto", "Product2.Name", "Product Name")
            )
            opportunity_name = _str(
                _cell(row, headers, "Nombre de la oportunidad", "Oportunidad", "Opportunity.Name", "Name")
            )
            if product_name is None and opportunity_name is None:
                continue

            close_date = _parse_date_text(
                _cell(row, headers, "Fecha de cierre", "Fecha cierre", "Close Date", "CloseDate")
            )
            if close_date is None or not product_name:
                continue

            unit_price = _decimal(_cell(row, headers, "Precio de venta", "Precio", "UnitPrice", "Sales Price")) or Decimal("0")
            quantity = _decimal(_cell(row, headers, "Cantidad", "Quantity")) or Decimal("1")
            account_name = _str(_cell(row, headers, "Nombre de la cuenta", "Cuenta", "Account.Name", "Account"))
            opp_key = f"{opportunity_name or ''}-{account_name or ''}-{close_date}"
            sf_opportunity_id = _str(_cell(row, headers, "Opportunity.Id", "Opportunity ID", "Id oportunidad"))
            if not sf_opportunity_id:
                sf_opportunity_id = "EXCEL_" + hashlib.md5(opp_key.encode()).hexdigest()[:12].upper()

            line_key = f"{opp_key}-{product_name}-{unit_price}-{quantity}"
            sf_line_item_id = _str(_cell(row, headers, "Id", "Line Item ID", "OpportunityLineItem.Id"))
            if not sf_line_item_id:
                sf_line_item_id = "EXCL_" + hashlib.md5(line_key.encode()).hexdigest()[:13].upper()

            licence_period = _decimal(
                _cell(row, headers, "Licence period (months)", "Meses", "Licence_Period_Months__c")
            )

            yield {
                "sf_opportunity_id": sf_opportunity_id,
                "sf_line_item_id": sf_line_item_id,
                "opportunity_owner": _str(_cell(row, headers, "Propietario de oportunidad", "Propietario", "Owner.Name")) or "Unknown",
                "account_name": account_name,
                "opportunity_name": opportunity_name,
                "opportunity_type": _str(_cell(row, headers, "Tipo", "Opportunity.Type")),
                "channel_type": _str(_cell(row, headers, "Tipo de oportunidad", "LeadSource", "Canal")),
                "opportunity_amount": _decimal(_cell(row, headers, "Importe", "Amount")),
                "close_date": close_date,
                "product_name": product_name,
                "unit_price": unit_price,
                "subscription_start_date": _parse_date_text(_cell(row, headers, "Subscription Start Date", "Inicio", "ServiceDate")),
                "subscription_end_date": _parse_date_text(_cell(row, headers, "Subscription End Date", "Fin", "EndDate")),
                "licence_period_months": int(licence_period) if licence_period is not None else None,
                "business_line": _str(_cell(row, headers, "Línea de negocio", "Linea de negocio", "Product2.Family")),
                "quantity": quantity,
                "product_code": _str(_cell(row, headers, "Product", "ProductCode", "Código producto", "Codigo")),
            }
            continue

        if row[9] is None and row[2] is None:
            continue

        close_date = _parse_date_text(row[6])
        if close_date is None:
            continue

        product_name = _str(row[9])
        if not product_name:
            continue

        unit_price = _decimal(row[10]) or Decimal("0")
        quantity = _decimal(row[15]) or Decimal("1")

        opp_key = f"{_str(row[2]) or ''}-{close_date}"
        sf_opportunity_id = "EXCEL_" + hashlib.md5(opp_key.encode()).hexdigest()[:12].upper()
        line_key = f"{opp_key}-{product_name}-{unit_price}-{quantity}"
        sf_line_item_id = "EXCL_" + hashlib.md5(line_key.encode()).hexdigest()[:13].upper()

        yield {
            "sf_opportunity_id": sf_opportunity_id,
            "sf_line_item_id": sf_line_item_id,
            "opportunity_owner": _str(row[0]) or "Unknown",
            "account_name": _str(row[1]),
            "opportunity_name": _str(row[2]),
            "opportunity_type": _str(row[3]),
            "channel_type": _str(row[4]),
            "opportunity_amount": _decimal(row[5]),
            "close_date": close_date,
            "product_name": product_name,
            "unit_price": unit_price,
            "subscription_start_date": _parse_date_text(row[11]),
            "subscription_end_date": _parse_date_text(row[12]),
            "licence_period_months": int(row[13]) if row[13] is not None else None,
            "business_line": _str(row[14]),
            "quantity": quantity,
            "product_code": _str(row[16]),
        }


def infer_product_classifications_from_rows(rows: list[dict]) -> dict:
    result = {}
    for row in rows:
        product_name = row["product_name"]
        if product_name in result:
            continue
        product_type = _infer_product_type(product_name, row.get("business_line"), row)
        if product_type:
            result[product_name] = {
                "product_code": row.get("product_code"),
                "business_line": row.get("business_line"),
                "category": "Salesforce raw",
                "subcategory": None,
                "product_type": product_type,
            }
    return result


def upsert_product_classifications(session: Session, products: dict):
    for name, info in products.items():
        # Skip compound keys — they exist only for in-memory lookup resolution
        if "|" in name:
            continue
        existing = session.query(ProductClassification).filter_by(product_name=name).first()
        if existing:
            existing.product_type = info["product_type"] or "UNCLASSIFIED"
            existing.product_code = info["product_code"]
            existing.category = info["category"]
            existing.subcategory = info["subcategory"]
            existing.business_line = info["business_line"]
        else:
            session.add(
                ProductClassification(
                    product_name=name,
                    product_type=info["product_type"] or "UNCLASSIFIED",
                    product_code=info["product_code"],
                    category=info["category"],
                    subcategory=info["subcategory"],
                    business_line=info["business_line"],
                )
            )
    session.flush()


def add_missing_product_placeholders(session: Session, products: dict, rows: list[dict]) -> dict:
    existing_names = {
        name for (name,) in session.query(ProductClassification.product_name).all()
    }
    merged = dict(products)
    for row in rows:
        product_name = row["product_name"]
        # Ignore compound keys in the check — plain name is the DB key
        if product_name in merged or product_name in existing_names:
            continue
        merged[product_name] = {
            "product_code": row.get("product_code"),
            "business_line": row.get("business_line"),
            "category": "Salesforce raw",
            "subcategory": None,
            "product_type": UNASSIGNED_MASTER_VALUE,
        }
    return merged


def upsert_consultant_countries(session: Session, countries: dict):
    for name, country in countries.items():
        existing = session.query(ConsultantCountry).filter_by(consultant_name=name).first()
        if existing:
            existing.country = country
        else:
            session.add(ConsultantCountry(consultant_name=name, country=country))
    session.flush()


def add_missing_consultant_placeholders(session: Session, countries: dict, rows: list[dict]) -> dict:
    existing_names = {
        name for (name,) in session.query(ConsultantCountry.consultant_name).all()
    }
    merged = dict(countries)
    for row in rows:
        consultant_name = row.get("opportunity_owner")
        if not consultant_name or consultant_name == "Unknown":
            continue
        if consultant_name in merged or consultant_name in existing_names:
            continue
        merged[consultant_name] = UNASSIGNED_MASTER_VALUE
    return merged


def create_snapshot(session: Session, *, triggered_by: str, notes: str) -> Snapshot:
    snap = Snapshot(
        id=uuid.uuid4(),
        sync_type="excel_import",
        triggered_by=triggered_by,
        status="processing",
        notes=notes,
    )
    session.add(snap)
    session.flush()
    return snap


def insert_raw_items(session: Session, snapshot_id, rows) -> list[RawOpportunityLineItem]:
    inserted = []
    for row in rows:
        item = RawOpportunityLineItem(
            id=uuid.uuid4(),
            snapshot_id=snapshot_id,
            sf_opportunity_id=row["sf_opportunity_id"],
            sf_line_item_id=row["sf_line_item_id"],
            opportunity_name=row["opportunity_name"],
            account_name=row["account_name"],
            opportunity_owner=row["opportunity_owner"],
            opportunity_type=row["opportunity_type"],
            channel_type=row["channel_type"],
            opportunity_amount=row["opportunity_amount"],
            close_date=row["close_date"],
            product_name=row["product_name"],
            product_code=row["product_code"],
            unit_price=row["unit_price"],
            quantity=row["quantity"],
            subscription_start_date=row["subscription_start_date"],
            subscription_end_date=row["subscription_end_date"],
            licence_period_months=row["licence_period_months"],
            business_line=row["business_line"],
        )
        session.add(item)
        inserted.append(item)
    session.flush()
    return inserted


def _all_months_in_snapshot(arr_snapshot) -> list[date]:
    months = set()
    for item in arr_snapshot.saas_items():
        m = item.start_month
        while m <= item.end_month_normalized:
            months.add(m)
            if m.month == 12:
                m = m.replace(year=m.year + 1, month=1)
            else:
                m = m.replace(month=m.month + 1)
    return sorted(months)


def run_calculation_and_store(session: Session, snapshot: Snapshot, raw_items):
    # Build lookup dict that supports (name, business_line) compound key resolution.
    # ARRCalculator still receives a flat {key: product_type} dict; compound keys
    # like "Usuarios|isEazy LMS" take precedence over plain "Usuarios" when the
    # raw item carries a business_line value.
    classifications = session.query(ProductClassification).all()
    products: dict[str, str] = {}
    for p in classifications:
        products[p.product_name] = p.product_type
        if p.business_line:
            products[f"{p.product_name}|{p.business_line}"] = p.product_type
    countries = {
        c.consultant_name: c.country for c in session.query(ConsultantCountry).all()
    }

    calc = ARRCalculator(products, countries)
    domain_items = []
    raw_by_id = {}
    for raw in raw_items:
        domain = RawLineItem(
            sf_opportunity_id=str(raw.sf_opportunity_id),
            sf_line_item_id=str(raw.sf_line_item_id),
            opportunity_name=raw.opportunity_name or "",
            account_name=raw.account_name or "",
            opportunity_owner=raw.opportunity_owner or "",
            opportunity_type=raw.opportunity_type or "",
            channel_type=raw.channel_type or "",
            close_date=raw.close_date,
            product_name=raw.product_name,
            unit_price=Decimal(str(raw.unit_price)),
            quantity=Decimal(str(raw.quantity)),
            subscription_start_date=raw.subscription_start_date,
            subscription_end_date=raw.subscription_end_date,
            licence_period_months=raw.licence_period_months,
            business_line=raw.business_line,
            opportunity_amount=Decimal(str(raw.opportunity_amount))
            if raw.opportunity_amount
            else None,
            product_code=raw.product_code,
        )
        domain_items.append(domain)
        raw_by_id[raw.sf_line_item_id] = raw.id

    arr_snapshot = calc.process_all(domain_items)

    for result in arr_snapshot.line_items:
        session.add(
            ARRLineItem(
                id=uuid.uuid4(),
                snapshot_id=snapshot.id,
                raw_line_item_id=raw_by_id[result.raw.sf_line_item_id],
                product_type=result.product_type,
                is_saas=result.is_saas,
                effective_start_date=result.effective_start_date,
                effective_end_date=result.effective_end_date,
                used_start_fallback=result.used_start_fallback,
                used_end_fallback=result.used_end_fallback,
                start_month=result.start_month,
                end_month_normalized=result.end_month_normalized,
                service_days=result.service_days,
                real_price=result.real_price,
                daily_price=result.daily_price,
                annualized_value=result.annualized_value,
                consultant_country=result.consultant_country,
                data_quality_flags=result.data_quality_flags,
            )
        )
    session.flush()

    all_months = _all_months_in_snapshot(arr_snapshot)
    monthly = calc.build_monthly_summary(arr_snapshot, all_months)
    for month_start, by_product in monthly.items():
        for product_type, arr_value in by_product.items():
            month_end = (
                month_start.replace(month=month_start.month % 12 + 1, day=1) - timedelta(days=1)
                if month_start.month < 12
                else month_start.replace(day=31)
            )
            count = sum(
                1
                for item in arr_snapshot.saas_items()
                if item.product_type == product_type
                and item.start_month <= month_end
                and item.end_month_normalized >= month_start
            )
            session.add(
                ARRMonthlySummary(
                    snapshot_id=snapshot.id,
                    month=month_start,
                    product_type=product_type,
                    arr_value=arr_value,
                    line_items_count=count,
                )
            )
    session.flush()

    for alert in arr_snapshot.alerts:
        session.add(
            SnapshotAlert(
                id=uuid.uuid4(),
                snapshot_id=snapshot.id,
                alert_type=alert["alert_type"],
                severity=alert["severity"],
                sf_opportunity_id=alert.get("sf_opportunity_id"),
                opportunity_name=alert.get("opportunity_name"),
                account_name=alert.get("account_name"),
                product_name=alert.get("product_name"),
                description=alert["description"],
            )
        )
    session.flush()

    quality = summarize_snapshot_quality(arr_snapshot)
    snapshot.status = "completed"
    snapshot.unclassified_products_count = quality["unclassified_products"]
    snapshot.alerts_count = quality["total_alerts"]
    snapshot.sf_records_fetched = len(raw_items)
    snapshot.sf_records_processed = len(arr_snapshot.line_items)
    return arr_snapshot


def import_excel_workbook(
    session: Session,
    wb,
    *,
    triggered_by: str = "api_excel_upload",
    notes: Optional[str] = None,
) -> ExcelImportSummary:
    t0 = datetime.now()

    products = load_product_classifications(wb)
    countries = load_consultant_countries(wb)
    rows = list(load_opos_rows(wb))

    if not rows:
        raise ExcelImportError("El Excel no contiene filas validas en 'Opos con Productos'.")

    try:
        products = add_missing_product_placeholders(session, products, rows)
        countries = add_missing_consultant_placeholders(session, countries, rows)

        upsert_product_classifications(session, products)
        upsert_consultant_countries(session, countries)

        snapshot = create_snapshot(
            session,
            triggered_by=triggered_by,
            notes=notes or "Importado manualmente desde Excel",
        )
        raw_items = insert_raw_items(session, snapshot.id, rows)
        run_calculation_and_store(session, snapshot, raw_items)

        snapshot.duration_seconds = round((datetime.now() - t0).total_seconds(), 2)
        session.commit()
        session.refresh(snapshot)
        return ExcelImportSummary(snapshot=snapshot, rows_read=len(rows))
    except Exception:
        session.rollback()
        raise


def import_excel_bytes(
    session: Session,
    file_bytes: bytes,
    *,
    triggered_by: str = "api_excel_upload",
    notes: Optional[str] = None,
) -> ExcelImportSummary:
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError("No se pudo abrir el fichero Excel. Usa un .xlsx valido.") from exc

    return import_excel_workbook(
        session,
        wb,
        triggered_by=triggered_by,
        notes=notes,
    )
